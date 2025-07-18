"""
Satranç Turnuva Etiketi ve Katılım Belgesi Oluşturucu

Bu uygulama, satranç turnuvalarında kullanılmak üzere madalya/kupa etiketleri
ve katılım belgeleri oluşturmak için geliştirilmiştir.

Özellikler:
- PDF ve Word formatında çıktı
- Çoklu kategori desteği
- Logo entegrasyonu
- Toplu katılım belgesi oluşturma
- Modern kullanıcı arayüzü

Geliştirici: Ertuğrul Kamil ŞAHİN
Satranç Hakemi
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter import font as tkfont
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
import textwrap
import datetime
import tempfile
import subprocess
from tkcalendar import Calendar
from PIL import Image, ImageTk, ImageOps, ImageEnhance
from reportlab.lib.colors import HexColor
import sys
from functools import partial
import logging
import smtplib
from email.mime.text import MIMEText
from openpyxl import load_workbook, Workbook
import shutil
import requests

def resource_path(relative_path):
    """PyInstaller ve normal çalışmada dosya yolu bulur."""
    base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
    return os.path.join(base_path, relative_path)

# BASE_DIR ve FONT_DIR sadeleştirilmiş tanım
BASE_DIR = getattr(sys, '_MEIPASS', os.path.dirname(__file__))
FONT_DIR = os.path.join(BASE_DIR, "ttf")

# Hata log dosyası ayarı
LOG_PATH = os.path.join(os.path.dirname(sys.argv[0]), "hata_kaydi.log")
logging.basicConfig(filename=LOG_PATH, level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')

def log_uncaught_exceptions(exctype, value, tb):
    import traceback
    logging.error("Beklenmeyen hata:", exc_info=(exctype, value, tb))
    # print("Beklenmeyen hata oluştu. Detaylar hata_kaydi.log dosyasına kaydedildi.")

sys.excepthook = log_uncaught_exceptions

PLAKA_ILLER = {
    1: "ADANA", 2: "ADIYAMAN", 3: "AFYONKARAHİSAR", 4: "AĞRI", 5: "AMASYA", 6: "ANKARA", 7: "ANTALYA", 8: "ARTVİN", 9: "AYDIN", 10: "BALIKESİR",
    11: "BİLECİK", 12: "BİNGÖL", 13: "BİTLİS", 14: "BOLU", 15: "BURDUR", 16: "BURSA", 17: "ÇANAKKALE", 18: "ÇANKIRI", 19: "ÇORUM", 20: "DENİZLİ",
    21: "DİYARBAKIR", 22: "EDİRNE", 23: "ELAZIĞ", 24: "ERZİNCAN", 25: "ERZURUM", 26: "ESKİŞEHİR", 27: "GAZİANTEP", 28: "GİRESUN", 29: "GÜMÜŞHANE", 30: "HAKKARİ",
    31: "HATAY", 32: "ISPARTA", 33: "MERSİN", 34: "İSTANBUL", 35: "İZMİR", 36: "KARS", 37: "KASTAMONU", 38: "KAYSERİ", 39: "KIRKLARELİ", 40: "KIRŞEHİR",
    41: "KOCAELİ", 42: "KONYA", 43: "KÜTAHYA", 44: "MALATYA", 45: "MANİSA", 46: "KAHRAMANMARAŞ", 47: "MARDİN", 48: "MUĞLA", 49: "MUŞ", 50: "NEVŞEHİR",
    51: "NİĞDE", 52: "ORDU", 53: "RİZE", 54: "SAKARYA", 55: "SAMSUN", 56: "SİİRT", 57: "SİNOP", 58: "SİVAS", 59: "TEKİRDAĞ", 60: "TOKAT",
    61: "TRABZON", 62: "TUNCELİ", 63: "ŞANLIURFA", 64: "UŞAK", 65: "VAN", 66: "YOZGAT", 67: "ZONGULDAK", 68: "AKSARAY", 69: "BAYBURT", 70: "KARAMAN",
    71: "KIRIKKALE", 72: "BATMAN", 73: "ŞIRNAK", 74: "BARTIN", 75: "ARDAHAN", 76: "IĞDIR", 77: "YALOVA", 78: "KARABÜK", 79: "KİLİS", 80: "OSMANİYE", 81: "DÜZCE"
}
def derece_metni(n):
    d = {
        1: "BİRİNCİSİ",
        2: "İKİNCİSİ",
        3: "ÜÇÜNCÜSÜ",
        4: "Dördüncüsü",
        5: "Beşincisi",
        6: "Altıncısı",
        7: "Yedincisi",
        8: "Sekizincisi",
        9: "Dokuzuncusu",
        10: "Onuncusu"
    }
    return d.get(n, f"{n}.")

def splash_screen_then_start(root, start_callback, splash_path, splash_time=2500):
    splash = tk.Toplevel(root)
    splash.overrideredirect(True)
    splash.geometry("600x400+400+200")
    try:
        img = Image.open(splash_path)
        img = img.resize((600, 400), Image.Resampling.LANCZOS)
        photo = ImageTk.PhotoImage(img)
        splash_label = tk.Label(splash, image=photo)
        splash_label.pack()
        global photo_ref
        photo_ref = photo
    except:
        label = tk.Label(splash, text="Başlangıç Görseli", font=HEADER_FONT)
        label.pack(expand=True, fill="both")
    def close_splash():
        splash.destroy()
        start_callback()
    splash.bind('<Escape>', lambda e: splash.destroy())
    root.after(splash_time, close_splash)

def parse_aralik(text):
    araliklar = []
    for parca in text.replace(" ","").split(","):
        if "-" in parca:
            try:
                bas, bit = map(int, parca.split("-"))
                if bas > 0 and bit >= bas:
                    araliklar.append((bas, bit))
            except:
                continue
        elif parca.isdigit():
            n = int(parca)
            if n > 0:
                araliklar.append((n, n))
    return araliklar

# --- Renk, Font ve Stil Sabitleri (Modern Pastel Palet) ---
PRIMARY_COLOR = "#2196f3"  # Modern mavi
SECONDARY_COLOR = "#43a047"  # Modern yeşil
DANGER_COLOR = "#e53935"  # Modern kırmızı
WARNING_COLOR = "#ffb300"  # Modern sarı
INFO_COLOR = "#00bcd4"  # Modern cyan
BG_COLOR = "#f5f7fa"  # Daha açık, modern arka plan
CARD_BG = "#ffffff"
BORDER_COLOR = "#e0e3e7"
APP_FONT_FAMILY = "Segoe UI"
HEADER_FONT = (APP_FONT_FAMILY, 16, "bold")
LABEL_FONT = (APP_FONT_FAMILY, 11)
ENTRY_FONT = (APP_FONT_FAMILY, 11)
BUTTON_FONT = (APP_FONT_FAMILY, 11)
SMALL_FONT = (APP_FONT_FAMILY, 9)
STEPPER_ACTIVE = "#2196f3"
STEPPER_DONE = "#43a047"
STEPPER_TODO = "#bdbdbd"
STEPPER_BG = "#e3f2fd"
ZEBRA_ROW = "#f0f4ff"

# Koyu tema için altyapı (isteğe bağlı)
DARK_BG = "#23272f"
DARK_CARD = "#2c313a"
DARK_TEXT = "#e0e3e7"
DARK_PRIMARY = "#90caf9"
DARK_SECONDARY = "#66bb6a"

# --- Uygulama Sabitleri ---
MAX_ODUL_PER_ROW = 5
DEFAULT_KATEGORI_SAYISI = 4
DEFAULT_MADALYA_W = 3.4
DEFAULT_MADALYA_H = 2.0
DEFAULT_KUPA_W = 6.0
DEFAULT_KUPA_H = 3.5
DEFAULT_FONT = "DejaVuSans"
DEFAULT_FONT_SIZE = 10

# --- EtiketUygulamasi ve diğer fonksiyonlar ---

class EtiketUygulamasi:
    def __init__(self, root):
        self.root = root
        self.root.title("Madalya/Kupa Etiketi-Katılım Belgesi Oluşturucu")
        self.root.geometry("900x600")
        self.root.minsize(700, 500)
        _icon_photos = []  # PNG ikon referanslarını tutmak için
        self.plaka_no = tk.StringVar()
        self.il_adi = tk.StringVar()
        self.turnuva_adi = tk.StringVar()
        self.kategori_sayisi = tk.IntVar(value=DEFAULT_KATEGORI_SAYISI)
        self.bas_tarih = tk.StringVar(value="Seçiniz")
        self.bit_tarih = tk.StringVar(value="Seçiniz")
        self.kategori_adlari_vars = [] if not hasattr(self, 'kategori_adlari_vars') or self.kategori_adlari_vars is None else self.kategori_adlari_vars
        self.odul_listeleri = [] if not hasattr(self, 'odul_listeleri') or self.odul_listeleri is None else self.odul_listeleri
        self.madalya_w = tk.StringVar(value=str(DEFAULT_MADALYA_W))
        self.madalya_h = tk.StringVar(value=str(DEFAULT_MADALYA_H))
        self.kupa_w = tk.StringVar(value=str(DEFAULT_KUPA_W))
        self.kupa_h = tk.StringVar(value=str(DEFAULT_KUPA_H))
        self.step = 0
        self.max_step = 3
        self.pdf_path = None
        self.preview_tempfile = None
        self.madalya_font = tk.StringVar(value=DEFAULT_FONT)
        self.kupa_font = tk.StringVar(value=DEFAULT_FONT)
        self.madalya_fontsize = tk.StringVar(value=str(DEFAULT_FONT_SIZE))
        self.kupa_fontsize = tk.StringVar(value=str(DEFAULT_FONT_SIZE))
        self.madalya_color = tk.StringVar(value="#000000")
        self.kupa_color = tk.StringVar(value="#000000")
        self.font_list = self.get_font_list()
        self.font_warning_shown = False
        self.tooltips = []  # Tooltip referansları
        self.stepper_logo_img = None
        # --- Logoları bellekte hazırla (sadece PIL.Image) ---
        # self.watermark_img = None
        # try:
        #     img = Image.open(resource_path("samsuntsflogo.png"))
        #     self.watermark_img = img.resize((350, 350), Image.Resampling.LANCZOS)
        # except Exception as e:
        #     print("Logo önbelleğe alınamadı:", e)
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TButton', font=BUTTON_FONT, background=PRIMARY_COLOR, foreground='white', borderwidth=0, focusthickness=3, focuscolor=PRIMARY_COLOR, padding=8)
        style.map('TButton', background=[('active', SECONDARY_COLOR)])
        style.configure('TLabel', background=BG_COLOR, font=LABEL_FONT)
        style.configure('Card.TFrame', background=CARD_BG, borderwidth=1, relief='ridge')
        style.configure('TEntry', font=ENTRY_FONT)
        style.configure('Treeview', font=ENTRY_FONT, rowheight=28, fieldbackground=CARD_BG, background=CARD_BG)
        style.map('Treeview', background=[('selected', PRIMARY_COLOR)])
        self.sol_logo_path = tk.StringVar()
        self.sag_logo_path = tk.StringVar()
        self.sol_logo_trace_id = None
        self.sag_logo_trace_id = None
        self.root.bind('<Right>', lambda e: self.next_step())
        self.root.bind('<Left>', lambda e: self.prev_step())
        self.root.bind('<Return>', lambda e: self.next_step())
        menubar = tk.Menu(self.root)
        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="Çıkış", command=self.root.quit)
        menubar.add_cascade(label="Dosya", menu=filemenu)
        helpmenu = tk.Menu(menubar, tearoff=0)
        # helpmenu.add_command(label="Kullanım Kılavuzu", command=self.show_help)  # KALDIRILDI
        helpmenu.add_command(label="Kullanım Kılavuzunu İndir", command=self.download_kilavuz)
        helpmenu.add_command(label="Geri Bildirimde Bulun", command=self.geri_bildirim_popup)
        menubar.add_cascade(label="Yardım", menu=helpmenu)
        aboutmenu = tk.Menu(menubar, tearoff=0)
        aboutmenu.add_command(label="Hakkında", command=self.show_about)
        menubar.add_cascade(label="Hakkında", menu=aboutmenu)
        updatemenu = tk.Menu(menubar, tearoff=0)
        updatemenu.add_command(label="Güncellemeleri Kontrol Et", command=self.guncelleme_kontrol_et)
        # updatemenu.add_command(label="Son Değişiklikler", command=self.show_changelog)  # KALDIRILDI
        menubar.add_cascade(label="Güncelleme", menu=updatemenu)
        self.root.config(menu=menubar)
        # --- Header grid: isim ve mail sağa yaslı ---
        info_frame = tk.Frame(self.root, bg=BG_COLOR)
        info_frame.pack(fill="x", side="top", padx=0, pady=(0, 0))
        version_label = tk.Label(info_frame, text=f"Sürüm: {APP_VERSION}", font=SMALL_FONT, bg=BG_COLOR, fg="#888", anchor="w")
        version_label.grid(row=0, column=0, sticky="w", padx=(8,0))
        right_frame = tk.Frame(info_frame, bg=BG_COLOR)
        author_label = tk.Label(right_frame, text="Ertuğrul Kamil ŞAHİN / Samsun", font=SMALL_FONT, bg=BG_COLOR, fg="#1976d2", anchor="e", justify="right")
        author_label.pack(side="top", anchor="e")
        email_label = tk.Label(right_frame, text="ertugrul.yazilim@gmail.com", font=SMALL_FONT, bg=BG_COLOR, fg="#1976d2", anchor="e", justify="right")
        email_label.pack(side="top", anchor="e")
        right_frame.grid(row=0, column=2, rowspan=2, sticky="e", padx=(0,8), pady=2)
        info_frame.grid_columnconfigure(0, weight=1)
        info_frame.grid_columnconfigure(1, weight=0)
        info_frame.grid_columnconfigure(2, weight=0)
     
        # --- Sürüm yazısı en sola alındı ---
        # --- Sağda metinler ve logo yan yana ---
        right_frame = tk.Frame(info_frame, bg=BG_COLOR)
        author_label = tk.Label(right_frame, text="Ertuğrul Kamil ŞAHİN / Samsun", font=SMALL_FONT, bg=BG_COLOR, fg="#1976d2", anchor="e")
        author_label.pack(side="top", anchor="e")
        email_label = tk.Label(right_frame, text="ertugrul.yazilim@gmail.com", font=SMALL_FONT, bg=BG_COLOR, fg="#1976d2", anchor="e")
        email_label.pack(side="top", anchor="e")
        # --- Sağda önce metin, sonra logo ---
        
        info_frame.grid_columnconfigure(0, weight=1)
        info_frame.grid_columnconfigure(1, weight=1)
        info_frame.grid_columnconfigure(2, weight=1)
        info_frame.grid_columnconfigure(3, weight=0)
        # Sağ alt logo kodu kaldırıldı, build_step'e taşınacak
        self.main_content = tk.Frame(self.root, bg=BG_COLOR)
        self.main_content.pack(expand=True, fill="both")
        self.last_il = ""
        self.last_turnuva = ""
        self.last_tarih = ""
        self.last_mesaj = ""
        self.last_logo_sol = ""
        self.last_logo_sag = ""
        self.last_direktor = ""
        self.last_bashakem = ""
        self.build_step()
        self.root.after(1000, self.guncelleme_kontrol_et)  # Açılışta otomatik güncelleme kontrolü
        # Tüm uygulama için varsayılan fontu ayarla
        default_font = tkfont.nametofont("TkDefaultFont")
        default_font.config(family=APP_FONT_FAMILY)
        text_font = tkfont.nametofont("TkTextFont")
        text_font.config(family=APP_FONT_FAMILY)
        fixed_font = tkfont.nametofont("TkFixedFont")
        fixed_font.config(family=APP_FONT_FAMILY)
        menu_font = tkfont.nametofont("TkMenuFont")
        menu_font.config(family=APP_FONT_FAMILY)
        heading_font = tkfont.nametofont("TkHeadingFont")
        heading_font.config(family=APP_FONT_FAMILY)
        caption_font = tkfont.nametofont("TkCaptionFont")
        caption_font.config(family=APP_FONT_FAMILY)
        # Stepper için logo bellekte hazırla
        self.stepper_logo_img = None
        try:
            img = Image.open(resource_path("samsuntsflogo.png"))
            self.stepper_logo_img = img.resize((56, 56), Image.Resampling.LANCZOS)
        except Exception as e:
            print("Stepper logo yüklenemedi:", e)

    def get_font_list(self):
        font_dir = FONT_DIR
        font_files = [f for f in os.listdir(font_dir) if f.endswith(".ttf")]
        font_names = []
        for f in font_files:
            name = os.path.splitext(f)[0]
            font_names.append(name)
            try:
                pdfmetrics.registerFont(TTFont(name, os.path.join(font_dir, f)))
            except Exception as e:
                print(f"Font kaydı başarısız: {name} - {e}")
        return font_names

    def clear_tooltips(self):
        for tip in self.tooltips:
            try:
                tip.destroy()
            except:
                pass
        self.tooltips = []

    def add_tooltip(self, widget, text):
        def on_enter(event):
            x, y = event.x_root + 20, event.y_root + 10
            tip = tk.Toplevel(widget)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{x}+{y}")
            label = tk.Label(tip, text=text, justify='left', background='#333', foreground='white', relief='solid', borderwidth=1, font=SMALL_FONT, padx=6, pady=3)
            label.pack()
            self.tooltips.append(tip)
        def on_leave(event):
            self.clear_tooltips()
        widget.bind('<Enter>', on_enter)
        widget.bind('<Leave>', on_leave)

    def build_step(self):
        for widget in self.main_content.winfo_children():
            widget.destroy()
        self.clear_tooltips()
        ana_frame = tk.Frame(self.main_content, bg=BG_COLOR)
        ana_frame.pack(expand=True, fill="both")

        # Stepper ve nav butonları
        top_frame = tk.Frame(ana_frame, bg=BG_COLOR)
        top_frame.pack(side=tk.TOP, fill="x")
        self._build_stepper(parent=top_frame)
        self._add_nav_buttons(parent=top_frame)

        # Adım içeriği
        content_frame = tk.Frame(ana_frame, bg=BG_COLOR)
        content_frame.pack(expand=True, fill="both")
        if self.step == 0:
            self._build_step0(parent=content_frame)
        elif self.step == 1:
            self._build_step1(parent=content_frame)
        elif self.step == 2:
            self._build_step2(parent=content_frame)
        elif self.step == 3:
            self._build_step3(parent=content_frame)

    def _build_stepper(self, parent):
        stepper_frame = tk.Frame(parent, bg=BG_COLOR, bd=0, highlightthickness=0)
        stepper_frame.pack(fill="x", pady=(2, 0), ipady=1)
        step_names = [
            ("Turnuva", "📝"),
            ("Kategoriler", "📋"),
            ("Etiket Ayarı", "🔤"),
            ("PDF", "📄")
        ]
        num_steps = len(step_names)
        for i, (name, icon_char) in enumerate(step_names):
            is_active = (i == self.step)
            is_done = (i < self.step)
            color = STEPPER_ACTIVE if is_active else (STEPPER_DONE if is_done else STEPPER_TODO)
            font_color = PRIMARY_COLOR if is_active else ("#b0bec5" if not is_done else SECONDARY_COLOR)
            step_btn = tk.Frame(stepper_frame, bg=BG_COLOR)
            step_btn.grid(row=0, column=2*i, sticky="nsew")
            stepper_frame.grid_columnconfigure(2*i, weight=1, uniform="step")
            icon_label = tk.Label(step_btn, text=icon_char, font=("Segoe UI Emoji", 20), fg=color, bg=BG_COLOR, cursor="hand2")
            icon_label.pack(pady=(0,2))
            text_label = tk.Label(step_btn, text=name, font=("Poppins", 12, "bold"), fg=font_color, bg=BG_COLOR, cursor="hand2")
            text_label.pack()
            if is_active:
                text_label.config(font=("Poppins", 12, "bold", "underline"))
            def goto_step(idx, *a):
                self.step = idx
                self.build_step()
            icon_label.bind("<Button-1>", partial(goto_step, i))
            text_label.bind("<Button-1>", partial(goto_step, i))
            def on_enter(e, lbl=text_label):
                lbl.config(fg=PRIMARY_COLOR, font=("Poppins", 12, "bold", "underline"))
            def on_leave(e, lbl=text_label, active=is_active):
                lbl.config(fg=PRIMARY_COLOR if active else ("#b0bec5" if not is_done else SECONDARY_COLOR), font=("Poppins", 12, "bold", "underline" if active else "normal"))
            text_label.bind("<Enter>", on_enter)
            text_label.bind("<Leave>", on_leave)
            icon_label.bind("<Enter>", on_enter)
            icon_label.bind("<Leave>", on_leave)
            # Ok simgesi (unicode)
            if i < num_steps-1:
                arrow_label = tk.Label(stepper_frame, text="➡️", font=("Segoe UI Emoji", 18), fg="#b0bec5", bg=BG_COLOR)
                arrow_label.grid(row=0, column=2*i+1, sticky="nsew")
                stepper_frame.grid_columnconfigure(2*i+1, weight=0)
        # Stepper'ın en sağına logo ekle (her zaman görünür)
        def add_stepper_logo():
            try:
                if self.stepper_logo_img:
                    stepper_logo_tk = ImageTk.PhotoImage(self.stepper_logo_img)
                    logo_label = tk.Label(stepper_frame, image=stepper_logo_tk, bg=BG_COLOR, borderwidth=0)
                    setattr(logo_label, 'image', stepper_logo_tk)  # Referans kaybolmasın
                    logo_label.grid(row=0, column=2*num_steps, sticky="e", padx=(8, 12), pady=2)
            except Exception as e:
                print("Stepper logo yüklenemedi:", e)
            stepper_frame.grid_columnconfigure(2*num_steps, weight=0)
        add_stepper_logo()
        stepper_frame.after_idle(add_stepper_logo)

    def _add_nav_buttons(self, parent):
        nav_frame = tk.Frame(parent, bg="#f7f7f7")
        nav_frame.pack(side=tk.TOP, fill="x", pady=1)
        inner_frame = tk.Frame(nav_frame, bg="#f7f7f7")
        inner_frame.pack(anchor="center")
        if self.step > 0 and self.step < 3:
            self.styled_button(inner_frame, text="← Geri", command=self.prev_step).pack(side=tk.LEFT, padx=14, pady=6)
        if self.step < 3:
            self.styled_button(inner_frame, text="İleri →", command=self.next_step).pack(side=tk.LEFT, padx=14, pady=6)
        if self.step == 3:
            self.styled_button(inner_frame, text="← Geri", command=self.prev_step).pack(side=tk.LEFT, padx=14, pady=6)

    def _build_step0(self, parent):
        main_frame = tk.Frame(parent, bg=BG_COLOR, bd=0, highlightthickness=0)
        main_frame.pack(fill="both", expand=True)
        card = tk.Frame(main_frame, bg=CARD_BG, bd=0, highlightthickness=0)
        card.pack(fill="both", expand=True, padx=10, pady=10)
        card.config(highlightbackground="#e0e3e7", highlightcolor="#e0e3e7", highlightthickness=2)
        tk.Label(card, text="Turnuva Bilgileri", font=HEADER_FONT, bg=BG_COLOR).pack(pady=6)
        plaka_frame = tk.LabelFrame(card, text="İl Bilgisi", font=LABEL_FONT, bg=CARD_BG, bd=2, relief="groove", fg=PRIMARY_COLOR)
        plaka_frame.pack(fill="x", pady=2)
        tk.Label(plaka_frame, text="Turnuvayı Düzenleyen İl (Plaka No):", font=LABEL_FONT, bg=CARD_BG).pack(side=tk.LEFT)
        plaka_entry = tk.Entry(plaka_frame, textvariable=self.plaka_no, width=5, font=ENTRY_FONT, bg="#f0f4ff", relief="solid", bd=1)
        plaka_entry.pack(side=tk.LEFT, padx=5)
        self.add_tooltip(plaka_entry, "Plaka numarası giriniz (örn: 34)")
        il_label = tk.Label(plaka_frame, text="", font=LABEL_FONT, fg=PRIMARY_COLOR, bg=CARD_BG)
        il_label.pack(side=tk.LEFT, padx=10)
        def plaka_guncelle(*args):
            try:
                plaka = int(self.plaka_no.get())
                il = PLAKA_ILLER.get(plaka, "")
                self.il_adi.set(il)
                il_label.config(text=il)
            except:
                self.il_adi.set("")
                il_label.config(text="")
        self.plaka_no.trace_add('write', plaka_guncelle)
        turnuva_frame = tk.LabelFrame(card, text="Turnuva Adı", font=LABEL_FONT, bg=CARD_BG, bd=2, relief="groove", fg=PRIMARY_COLOR)
        turnuva_frame.pack(fill="x", pady=2)
        turnuva_entry = tk.Entry(turnuva_frame, textvariable=self.turnuva_adi, width=50, font=ENTRY_FONT, bg="#f0f4ff", relief="solid", bd=1)
        turnuva_entry.pack(side=tk.LEFT, fill="x", expand=True, padx=2, pady=4)
        self.add_tooltip(turnuva_entry, "Turnuva adını yazınız")
        tk.Label(turnuva_frame, text="(SATRANÇ TURNUVASI ibaresi otomatik eklenmektedir.)", font=SMALL_FONT, bg=CARD_BG).pack(side=tk.LEFT, padx=8)
        tarih_frame = tk.LabelFrame(card, text="Tarih Bilgisi", font=LABEL_FONT, bg=CARD_BG, bd=2, relief="groove", fg=PRIMARY_COLOR)
        tarih_frame.pack(anchor="w", pady=(2, 0), fill="x")
        tk.Label(tarih_frame, text="Başlangıç Tarihi:", font=LABEL_FONT, bg=CARD_BG).pack(side=tk.LEFT)
        bas_date_btn = self.styled_button(tarih_frame, textvariable=self.bas_tarih, width=12, command=lambda: TarihSecici(self.root, self.bas_tarih, self.bas_tarih.get()))
        bas_date_btn.pack(side=tk.LEFT, padx=5)
        tk.Label(tarih_frame, text="-", font=LABEL_FONT, bg=CARD_BG).pack(side=tk.LEFT, padx=5)
        tk.Label(tarih_frame, text="Bitiş Tarihi:", font=LABEL_FONT, bg=CARD_BG).pack(side=tk.LEFT)
        bit_date_btn = self.styled_button(tarih_frame, textvariable=self.bit_tarih, width=12, command=lambda: TarihSecici(self.root, self.bit_tarih, self.bit_tarih.get()))
        bit_date_btn.pack(side=tk.LEFT, padx=5)
        tk.Label(card, text="Kategori Sayısı:", font=LABEL_FONT, bg=BG_COLOR).pack(anchor="w", pady=(2, 0))
        kategori_sayisi_frame = tk.Frame(card, bg=BG_COLOR)
        kategori_sayisi_frame.pack(anchor="w", pady=1)
        self.kategori_sayisi_option = tk.StringVar(value=str(self.kategori_sayisi.get()))
        kategori_menu = ttk.Combobox(kategori_sayisi_frame, textvariable=self.kategori_sayisi_option, values=[str(i) for i in range(1, 21)], font=ENTRY_FONT, width=6, state="readonly")
        kategori_menu.pack(side=tk.LEFT)
        self.add_tooltip(kategori_menu, "Kategori sayısı 1-20 arası olmalı")
        def sync_kategori_sayisi_var(*args):
            if self.kategori_sayisi_option.get().isdigit():
                self.kategori_sayisi.set(int(self.kategori_sayisi_option.get()))
        def sync_kategori_option_var(*args):
            self.kategori_sayisi_option.set(str(self.kategori_sayisi.get()))
        self.kategori_sayisi.trace_add('write', sync_kategori_option_var)
        self.kategori_sayisi_option.trace_add('write', sync_kategori_sayisi_var)

    def _build_step1(self, parent):
        main_frame = tk.Frame(parent, bg=BG_COLOR, bd=0, highlightthickness=0)
        main_frame.pack(fill="both", expand=True)
        card = tk.Frame(main_frame, bg=CARD_BG, bd=0, highlightthickness=0)
        card.pack(fill="both", expand=True, padx=20, pady=20)
        card.config(highlightbackground="#e0e3e7", highlightcolor="#e0e3e7", highlightthickness=2)
        tk.Label(card, text="Kategori Ödülleri (Tablo)", font=HEADER_FONT, bg="#f7f7f7").pack(pady=12)
        tk.Label(card, text="Kategori Adını Yazınız. 'KATEGORİSİ' İbaresi Etikette Otomatik Eklenecektir.", font=LABEL_FONT, bg="#f7f7f7").pack(pady=(0, 10))
        columns = ["S.NO", "KATEGORİ", "ÖDÜLLER"]
        # --- SCROLLABLE TABLE ---
        table_outer = tk.Frame(card, bg=BG_COLOR)
        table_outer.pack(fill="both", expand=True, pady=10)
        col_widths = [6, 32, 32]
        table_canvas = tk.Canvas(table_outer, bg=BG_COLOR, highlightthickness=0, bd=0, height=520)
        table_canvas.pack(side="left", fill="both", expand=True)
        scrollbar = tk.Scrollbar(table_outer, orient="vertical", command=table_canvas.yview)
        scrollbar.pack(side="right", fill="y")
        table_canvas.configure(yscrollcommand=scrollbar.set)
        table_inner = tk.Frame(table_canvas, bg=BG_COLOR)
        table_window = table_canvas.create_window((0,0), window=table_inner, anchor="nw")
        def on_configure(event):
            table_canvas.configure(scrollregion=table_canvas.bbox("all"))
            table_canvas.itemconfig(table_window, width=table_canvas.winfo_width())
        table_inner.bind("<Configure>", on_configure)
        def resize_canvas(event):
            table_canvas.itemconfig(table_window, width=table_canvas.winfo_width())
        table_canvas.bind("<Configure>", resize_canvas)
        table_frame = table_inner
        # --- BAŞLIK SATIRINI DA TABLE_FRAME'E EKLE ---
        for col, (h, w) in enumerate(zip(columns, col_widths)):
            tk.Label(table_frame, text=h, font=("Segoe UI", 12, "bold"), borderwidth=1, relief="solid", width=w, bg="#e3f2fd").grid(row=0, column=col, sticky="nsew", ipady=6)
            table_frame.grid_columnconfigure(col, weight=1)
        toplam = self.kategori_sayisi.get()
        while len(self.kategori_adlari_vars) < toplam:
            self.kategori_adlari_vars.append(tk.StringVar())
        while len(self.kategori_adlari_vars) > toplam:
            self.kategori_adlari_vars.pop()
        while len(self.odul_listeleri) < toplam:
            self.odul_listeleri.append([])
        while len(self.odul_listeleri) > toplam:
            self.odul_listeleri.pop()
        for i in range(self.kategori_sayisi.get()):
            tk.Label(table_frame, text=str(i+1), borderwidth=1, relief="solid", width=col_widths[0], font=("Segoe UI", 11), bg=ZEBRA_ROW if i%2==0 else CARD_BG).grid(row=i+1, column=0, sticky="nsew", ipady=4)
            entry = tk.Entry(table_frame, textvariable=self.kategori_adlari_vars[i], borderwidth=1, relief="solid", font=("Segoe UI", 11), bg=ZEBRA_ROW if i%2==0 else CARD_BG, width=col_widths[1])
            entry.grid(row=i+1, column=1, sticky="nsew", ipady=4)
            table_frame.grid_columnconfigure(1, weight=1)
            def update_tooltip(var=self.kategori_adlari_vars[i], widget=entry):
                val = var.get()
                if len(val) > 20:
                    self.add_tooltip(widget, val)
                else:
                    self.add_tooltip(widget, "Kategori adını giriniz")
            self.kategori_adlari_vars[i].trace_add('write', lambda *a, var=self.kategori_adlari_vars[i], widget=entry: update_tooltip(var, widget))
            update_tooltip(self.kategori_adlari_vars[i], entry)
            odul_frame = tk.Frame(table_frame, borderwidth=1, relief="solid", bg=ZEBRA_ROW if i%2==0 else CARD_BG)
            odul_frame.grid(row=i+1, column=2, sticky="nsew")
            table_frame.grid_columnconfigure(2, weight=1)
            max_per_row = 4
            odul_list = self.odul_listeleri[i]
            for odul_idx, odul in enumerate(odul_list):
                row = odul_idx // max_per_row
                col = odul_idx % max_per_row
                # Kısa ödül metni
                if odul['tur'] == 'Katılım Madalyası':
                    odul_text = f"🎖️ Katılım [{odul['adet']}]"
                    tooltip_text = f"Katılım Madalyası (adet: {odul['adet']})"
                elif odul['tur'] == 'En İyi Kadın':
                    odul_text = f"👩‍🦰 En İyi Kadın [{odul['baslangic']}-{odul['bitis']}]"
                    tooltip_text = f"En İyi Kadın [{odul['baslangic']}-{odul['bitis']}]"
                elif odul['tur'] == 'Kupa':
                    odul_text = f"🏆 Kupa [{odul['baslangic']}-{odul['bitis']}]"
                    tooltip_text = f"Kupa [{odul['baslangic']}-{odul['bitis']}]"
                elif odul['tur'] == 'Madalya':
                    odul_text = f"🏅 Madalya [{odul['baslangic']}-{odul['bitis']}]"
                    tooltip_text = f"Madalya [{odul['baslangic']}-{odul['bitis']}]"
                elif odul['tur'] == 'Diğer':
                    if 'baslangic' in odul and 'bitis' in odul:
                        odul_text = f"⭐ {odul['metin']} [{odul['baslangic']}-{odul['bitis']}]"
                        tooltip_text = f"{odul['metin']} [{odul['baslangic']}-{odul['bitis']}]"
                    else:
                        odul_text = f"⭐ {odul['metin']}"
                        tooltip_text = odul['metin']
                else:
                    odul_text = odul['tur']
                    tooltip_text = odul['tur']
                odul_cell = tk.Frame(odul_frame, bg=ZEBRA_ROW if i%2==0 else CARD_BG)
                lbl = tk.Label(odul_cell, text=odul_text, font=("Segoe UI", 9, "bold"), fg=SECONDARY_COLOR, bg=odul_cell["bg"])
                lbl.pack()
                self.add_tooltip(lbl, tooltip_text)
                btns = tk.Frame(odul_cell, bg=odul_cell["bg"])
                btns.pack()
                self.styled_button(btns, text="Düzenle", font=("Segoe UI", 8), bg=WARNING_COLOR, activebackground="#ff9800", command=lambda idx=i, oidx=odul_idx: self.odul_duzenle_popup(idx, oidx), padx=6, pady=2).pack(side=tk.LEFT, padx=1)
                self.styled_button(btns, text="Kaldır", font=("Segoe UI", 8), bg=DANGER_COLOR, activebackground="#b71c1c", command=lambda idx=i, oidx=odul_idx: self.odul_kaldir(idx, oidx), padx=6, pady=2).pack(side=tk.LEFT, padx=1)
                odul_cell.grid(row=row, column=col, padx=2, pady=2)
            # Ödül Ekle butonu en sağa gelsin, alt satıra inmesin
            ekle_col = len(odul_list) % max_per_row
            ekle_row = len(odul_list) // max_per_row
            self.styled_button(odul_frame, text="Ödül Ekle", bg=INFO_COLOR, activebackground="#0277bd", command=lambda idx=i: self.odul_ekle_popup(idx), font=("Segoe UI", 9), padx=8, pady=3).grid(row=ekle_row, column=ekle_col, padx=2, pady=2, sticky="w")
            # Mouse tekerleği ile scroll desteği
            def _on_mousewheel(event):
                table_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            table_canvas.bind_all("<MouseWheel>", _on_mousewheel)

    def _build_step2(self, parent):
        def logo_sec(var, label, kaldir_btn):
            from tkinter import filedialog
            path = filedialog.askopenfilename(title="Logo Seç", filetypes=[("Resim Dosyası", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
            if path:
                var.set(path)
                if label is not None:
                    label.config(text=os.path.basename(path), fg="#388e3c")
                if kaldir_btn is not None:
                    kaldir_btn.grid()
            else:
                var.set("")
                if label is not None:
                    label.config(text="(Seçilmedi)", fg="#bdbdbd")
                if kaldir_btn is not None:
                    kaldir_btn.grid_remove()

        def logo_kaldir(var, label, kaldir_btn):
            var.set("")
            if label is not None:
                label.config(text="(Seçilmedi)", fg="#bdbdbd")
            if kaldir_btn is not None:
                kaldir_btn.grid_remove()
        main_frame = tk.Frame(parent, bg=BG_COLOR, bd=0, highlightthickness=0)
        main_frame.pack(fill="both", expand=True)
        card = tk.Frame(main_frame, bg=CARD_BG, bd=0, highlightthickness=0)
        card.pack(fill="both", expand=True, padx=20, pady=20)
        card.config(highlightbackground="#e0e3e7", highlightcolor="#e0e3e7", highlightthickness=2)
        tk.Label(card, text="Etiket Ölçüsü ve Yazı Tipi", font=HEADER_FONT, bg="#f7f7f7").pack(pady=12)
        for tip, w_var, h_var, font_var, size_var, color_var, label in [
            ("Madalya", self.madalya_w, self.madalya_h, self.madalya_font, self.madalya_fontsize, self.madalya_color, "Madalya etiketi"),
            ("Kupa", self.kupa_w, self.kupa_h, self.kupa_font, self.kupa_fontsize, self.kupa_color, "Kupa etiketi")
        ]:
            frame = tk.LabelFrame(card, text=label, font=BUTTON_FONT, bg="#f7f7f7", bd=2, relief="groove")
            frame.pack(pady=8, anchor="w", fill="x")
            for widget in frame.winfo_children():
                widget.destroy()
            # Grid: Genişlik | x | Yükseklik | cm | Yazı Tipi Seçiniz: [Buton] | Rengi Seçiniz: [Buton] | Önizleme
            tk.Label(frame, text="Genişlik:", font=LABEL_FONT, bg="#f7f7f7").grid(row=0, column=0, sticky="e", padx=4, pady=4)
            tk.Entry(frame, textvariable=w_var, width=5, font=ENTRY_FONT).grid(row=0, column=1, padx=2)
            tk.Label(frame, text="x", bg="#f7f7f7").grid(row=0, column=2)
            tk.Label(frame, text="Yükseklik:", font=LABEL_FONT, bg="#f7f7f7").grid(row=0, column=3, sticky="e", padx=2)
            tk.Entry(frame, textvariable=h_var, width=5, font=ENTRY_FONT).grid(row=0, column=4, padx=2)
            tk.Label(frame, text="cm", bg="#f7f7f7").grid(row=0, column=5, padx=2)
            # Yazı tipi seç
            tk.Label(frame, text="Yazı Tipi Seçiniz:", font=LABEL_FONT, bg="#f7f7f7").grid(row=0, column=6, sticky="e", padx=(10,2))
            font_btn = self.styled_button(frame, text=font_var.get(), bg="#222", fg="white", padx=4, pady=2, font=(font_var.get(), 10))
            font_btn.config(command=lambda v=font_var, b=font_btn, f=frame: self.font_popup(v, b, f))
            font_btn.grid(row=0, column=7, padx=2)
            def update_font_btn(*args, btn=font_btn, v=font_var):
                btn.config(text=v.get(), font=(v.get(), 10))
            font_var.trace_add('write', update_font_btn)
            # Renk seç
            tk.Label(frame, text="Rengi Seçiniz:", font=LABEL_FONT, bg="#f7f7f7").grid(row=0, column=8, sticky="e", padx=(10,2))
            renk_btn = self.styled_button(frame, text="Renk", bg=color_var.get(), padx=4, pady=2, font=LABEL_FONT)
            renk_btn.config(command=lambda v=color_var, b=renk_btn: self.pick_color(v, b))
            renk_btn.grid(row=0, column=9, padx=2)
            color_var.trace_add('write', lambda *args, v=color_var, b=renk_btn: self.update_btn_color(v, b))
            # Önizleme kutusu
            preview_lbl = tk.Label(frame, text="Önizleme", font=(font_var.get(), int(size_var.get())), fg=color_var.get(), bg="#fff", relief="solid", bd=1, width=14, anchor="center")
            preview_lbl.grid(row=0, column=10, padx=(16,4), pady=2)
            def update_preview(*args, lbl=preview_lbl, f=font_var, s=size_var, c=color_var):
                try:
                    lbl.config(font=(f.get(), int(s.get())), fg=c.get())
                except:
                    lbl.config(font=("Arial", 11), fg="#000")
            font_var.trace_add('write', update_preview)
            size_var.trace_add('write', update_preview)
            color_var.trace_add('write', update_preview)
            update_preview()
        logo_frame = tk.LabelFrame(card, text="Etiket Logoları (İsteğe Bağlı)", font=BUTTON_FONT, bg="#f7f7f7", bd=2, relief="groove")
        logo_frame.pack(pady=8, fill="x")
        # Sol logo satırı
        for widget in logo_frame.winfo_children():
            widget.destroy()
        def logo_preview_img(path):
            from PIL import Image, ImageTk, ImageOps
            if path and os.path.exists(path):
                try:
                    img = Image.open(path).convert("RGBA")
                    img = ImageOps.fit(img, (64, 64), Image.Resampling.LANCZOS)
                    return ImageTk.PhotoImage(img)
                except:
                    return None
            return None
        def update_logo_preview(var, img_label):
            path = var.get()
            img = logo_preview_img(path)
            if img:
                img_label.config(image=img, text="")
                img_label.image = img
            else:
                img_label.config(image="", text="Yok", fg="#bbb", font=SMALL_FONT, width=8, height=4)
                img_label.image = None
        # Sol logo
        tk.Label(logo_frame, text="Sol Üst Logo:", font=LABEL_FONT, bg="#f7f7f7").grid(row=0, column=0, sticky="e", padx=4, pady=4)
        self.styled_button(logo_frame, text="Logo Seç", command=lambda: logo_sec(self.sol_logo_path, sol_logo_label, sol_logo_kaldir_btn), bg=INFO_COLOR, activebackground="#0277bd", padx=4, pady=2, font=LABEL_FONT).grid(row=0, column=1, padx=2)
        sol_logo_label = tk.Label(logo_frame, text="(Seçilmedi)", font=(SMALL_FONT[0], SMALL_FONT[1], "italic"), bg="#f7f7f7", fg="#bdbdbd")
        sol_logo_label.grid(row=0, column=2, padx=4)
        sol_logo_kaldir_btn = self.styled_button(logo_frame, text="Kaldır", command=lambda: logo_kaldir(self.sol_logo_path, sol_logo_label, sol_logo_kaldir_btn), bg=DANGER_COLOR, activebackground="#b71c1c", padx=4, pady=2, font=LABEL_FONT)
        sol_logo_kaldir_btn.grid(row=0, column=3, padx=2)
        def sol_logo_update(*a):
            path = self.sol_logo_path.get()
            if path:
                sol_logo_label.config(text=os.path.basename(path), fg="#388e3c", font=(SMALL_FONT[0], SMALL_FONT[1], "bold"))
            else:
                sol_logo_label.config(text="(Seçilmedi)", fg="#bdbdbd", font=(SMALL_FONT[0], SMALL_FONT[1], "italic"))
        self.sol_logo_path.trace_add('write', lambda *a: sol_logo_update())
        sol_logo_update()
        # Sağ logo
        tk.Label(logo_frame, text="Sağ Üst Logo:", font=LABEL_FONT, bg="#f7f7f7").grid(row=1, column=0, sticky="e", padx=4, pady=4)
        self.styled_button(logo_frame, text="Logo Seç", command=lambda: logo_sec(self.sag_logo_path, sag_logo_label, sag_logo_kaldir_btn), bg=INFO_COLOR, activebackground="#0277bd", padx=4, pady=2, font=LABEL_FONT).grid(row=1, column=1, padx=2)
        sag_logo_label = tk.Label(logo_frame, text="(Seçilmedi)", font=(SMALL_FONT[0], SMALL_FONT[1], "italic"), bg="#f7f7f7", fg="#bdbdbd")
        sag_logo_label.grid(row=1, column=2, padx=4)
        sag_logo_kaldir_btn = self.styled_button(logo_frame, text="Kaldır", command=lambda: logo_kaldir(self.sag_logo_path, sag_logo_label, sag_logo_kaldir_btn), bg=DANGER_COLOR, activebackground="#b71c1c", padx=4, pady=2, font=LABEL_FONT)
        sag_logo_kaldir_btn.grid(row=1, column=3, padx=2)
        def sag_logo_update(*a):
            path = self.sag_logo_path.get()
            if path:
                sag_logo_label.config(text=os.path.basename(path), fg="#388e3c", font=(SMALL_FONT[0], SMALL_FONT[1], "bold"))
            else:
                sag_logo_label.config(text="(Seçilmedi)", fg="#bdbdbd", font=(SMALL_FONT[0], SMALL_FONT[1], "italic"))
        self.sag_logo_path.trace_add('write', lambda *a: sag_logo_update())
        sag_logo_update()

    def _build_step3(self, parent):
        from PIL import ImageEnhance
        container = tk.Frame(parent, bg=BG_COLOR)
        container.pack(expand=True, fill="both")
        card = tk.Frame(container, bg="#fff", bd=1, relief="solid")
        card.pack(expand=True)
        # Başlık
        tk.Label(card, text="PDF Önizleme ve Oluşturma", font=HEADER_FONT, bg="#fff").pack(pady=(24, 12))
        # Butonlar
        btn_frame = tk.Frame(card, bg="#fff")
        btn_frame.pack(pady=(0, 18))
        self.styled_button(btn_frame, text="PDF Önizle", bg="#00bcd4", command=self.pdf_onizle).pack(side="left", padx=8)
        self.styled_button(btn_frame, text="PDF Oluştur", bg="#388e3c", command=self.pdf_olustur).pack(side="left", padx=8)
        self.styled_button(btn_frame, text="İletişime Geç", command=self.geri_bildirim_popup, bg="#1976d2", activebackground="#115293").pack(side="left", padx=12)
        self.styled_button(btn_frame, text="Programı Kapat", bg="#d32f2f", command=self.root.quit).pack(side="left", padx=8)
        self.styled_button(btn_frame, text="Toplu Katılım Belgesi Oluştur", bg="#ffa000", command=self.toplu_katilim_popup).pack(side="left", padx=8)
        # Açıklama kutusu (kutu_frame)
        kutu_frame = tk.Frame(card, bg="#fff")
        kutu_frame.pack(pady=(8, 24), padx=8)
        desc_text = (
            "Bu uygulama, turnuvalarda görev yapan hakem arkadaşlarımın kupa ve madalya ödülleri için gerekli etiketleri daha pratik, hızlı ve sorunsuz bir şekilde hazırlayabilmesi amacıyla geliştirildi."
            "Turnuva temposunun yoğunluğunu bildiğim için, bu süreci sizin adınıza biraz daha kolaylaştırmak istedim."
            "\nSorularınız ve Katkılarınız için İletişime Geçmenizi Rica Ederim."
        )
        desc_label = tk.Label(
            kutu_frame,
            text=desc_text,
            font=("Segoe UI", 10),
            bg="#fff",
            wraplength=540,
            justify="center",
            padx=18,
            pady=14
        )
        desc_label.pack()
        # Kutunun hemen altında, ortalanmış logo
        try:
            img = Image.open(resource_path("samsuntsflogo.png")).convert("RGBA")
            img = img.resize((100, 100), Image.Resampling.LANCZOS)
            tk_img = ImageTk.PhotoImage(img)
            logo_label = tk.Label(card, image=tk_img, bg=BG_COLOR, bd=0)
            self.pdf_logo_img_bottom = tk_img  # Referans kaybolmasın
            logo_label.pack(pady=(0, 10))
        except Exception as e:
            print("Alt logo hatası:", e)
        # En altta geliştirici iletişim bilgisi
        dev_label = tk.Label(card, text="Geliştirici: Ertuğrul Kamil ŞAHİN / Samsun  |  ertugrul.yazilim@gmail.com", font=("Segoe UI", 9, "italic"), fg="#888", bg="#fff")
        dev_label.pack(pady=(8, 0))

    def geri_bildirim_mail_gonder(self, ad, cep, mesaj, mail):
        GONDEREN = 'ertugrul.yazilim@gmail.com'
        SIFRE = 'ugte ufja gtbe rbue'
        ALICI = 'ertugrul.yazilim@gmail.com'
        subject = 'Uygulama Geri Bildirim'
        body = f'Adınız- Soyadınız: {ad}\nE-posta adresiniz: {mail}\nCep: {cep}\nMesaj: {mesaj}'
        msg = MIMEText(body, 'plain', 'utf-8')
        msg['Subject'] = subject
        msg['From'] = GONDEREN
        msg['To'] = ALICI
        try:
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
                server.login(GONDEREN, SIFRE)
                server.sendmail(GONDEREN, ALICI, msg.as_string())
            return True
        except Exception as e:
            logging.error(f'Geri bildirim maili gönderilemedi: {e}')
            return False

    def geri_bildirim_popup(self):
        popup = tk.Toplevel(self.root)
        popup.title("Geri Bildirim Gönder")
        popup.geometry("480x520")
        popup.resizable(False, False)
        popup.configure(bg="#f5f7fa")
        popup.bind('<Escape>', lambda e: popup.destroy())
        label_font = ("Segoe UI", 10)
        entry_font = ("Segoe UI", 10)
        text_font = ("Segoe UI", 10)
        # --- Canvas+Scrollbar ile kartı kaydırılabilir yap ---
        canvas = tk.Canvas(popup, bg="#f5f7fa", highlightthickness=0, width=480, height=470)
        canvas.pack(side="top", fill="both", expand=True)
        scrollbar = tk.Scrollbar(popup, orient="vertical", command=canvas.yview)
        scrollbar.pack(side="right", fill="y")
        card = tk.Frame(canvas, bg="#fff", bd=0, highlightthickness=0, padx=12, pady=12)
        card_id = canvas.create_window((0,0), window=card, anchor="nw", width=480)
        def on_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        card.bind("<Configure>", on_configure)
        # --- Başlık ve açıklama ---
        tk.Label(card, text="Geri Bildirim", font=("Segoe UI", 16, "bold"), bg="#fff", fg="#1976d2").pack(pady=(0,2))
        tk.Label(card, text="Görüş, öneri veya hata bildirimi için aşağıdaki formu doldurabilirsiniz.", font=("Segoe UI", 8), bg="#fff", fg="#888", wraplength=400, justify="center").pack(pady=(0,12))
        # --- Alanlar ---
        alanlar = []
        def alan_satiri(label, var, zorunlu=False, placeholder=""):
            frame = tk.Frame(card, bg="#fff")
            frame.pack(fill="x", pady=(6,0))
            label_text = label + (" *" if zorunlu else "")
            lbl = tk.Label(frame, text=label_text, font=label_font, bg="#fff", fg="#d32f2f" if zorunlu else "#333")
            lbl.pack(anchor="w")
            entry = tk.Entry(frame, textvariable=var, font=entry_font, relief="solid", bd=1, bg="#f9f9fb", fg="#222")
            entry.pack(fill="x", ipady=5, padx=(0,0), pady=(0, 6))
            entry.insert(0, placeholder)
            alanlar.append((entry, zorunlu, lbl))
            return entry
        ad_var = tk.StringVar()
        mail_var = tk.StringVar()
        cep_var = tk.StringVar()
        mesaj_var = tk.StringVar()
        entry_ad = alan_satiri("Adınız-Soyadınız:", ad_var, zorunlu=True, placeholder="Adınızı girin")
        entry_mail = alan_satiri("E-posta Adresiniz:", mail_var, zorunlu=True, placeholder="E-posta adresinizi girin")
        entry_cep = alan_satiri("Cep Telefonu (isteğe bağlı):", cep_var, zorunlu=False, placeholder="5xx xxx xx xx")
        # Mesaj için ayrı Text kutusu
        frame_mesaj = tk.Frame(card, bg="#fff")
        frame_mesaj.pack(fill="x", pady=(6,0))
        lbl_mesaj = tk.Label(frame_mesaj, text="Geri Bildiriminizi Yazınız: *", font=label_font, bg="#fff", fg="#d32f2f")
        lbl_mesaj.pack(anchor="w")
        text_mesaj = tk.Text(frame_mesaj, font=text_font, height=7, width=50, relief="solid", bd=1, bg="#f9f9fb", fg="#222")
        text_mesaj.pack(fill="x", ipady=4, pady=(0,6))
        # Hata mesajı alanı
        hata_label = tk.Label(card, text="", font=("Segoe UI", 9), fg="#d32f2f", bg="#fff")
        hata_label.pack(pady=(4,0))
        # --- Butonlar kartın dışında, popup'ın en altında ---
        btn_frame = tk.Frame(popup, bg="#f5f7fa")
        btn_frame.pack(side="bottom", pady=(0,14))
        def on_enter(btn, color): btn.config(bg=color)
        def on_leave(btn, color): btn.config(bg=color)
        def show_toast(msg, color="#43a047"):
            toast = tk.Toplevel(popup)
            toast.overrideredirect(True)
            toast.geometry(f"300x40+{popup.winfo_rootx()+60}+{popup.winfo_rooty()+20}")
            toast.configure(bg=color)
            tk.Label(toast, text=msg, font=("Segoe UI", 11, "bold"), bg=color, fg="#fff").pack(expand=True, fill="both")
            toast.after(1800, toast.destroy)
        def gonder():
            ad = ad_var.get().strip()
            mail = mail_var.get().strip()
            cep = cep_var.get().strip()
            mesaj = text_mesaj.get("1.0", "end").strip()
            hata_label.config(text="")
            # Alan kontrolü ve görsel uyarı
            eksik = False
            for entry, zorunlu, lbl in alanlar:
                entry.config(highlightthickness=0, highlightbackground="#ccc")
                if zorunlu and not entry.get().strip():
                    entry.config(highlightthickness=2, highlightbackground="#d32f2f")
                    eksik = True
            if not ad or not mail or not mesaj or mesaj == "Geri bildiriminizi yazınız":
                hata_label.config(text="Lütfen zorunlu alanları doldurunuz.")
                return
            if self.geri_bildirim_mail_gonder(ad, cep, mesaj, mail):
                show_toast("Geri bildiriminiz gönderildi!", color="#43a047")
                popup.after(1800, popup.destroy)
            else:
                hata_label.config(text="Geri bildirim e-posta ile gönderilemedi. Lütfen daha sonra tekrar deneyin.")
        # Butonları ortala
        btn_gonder = tk.Button(btn_frame, text="📧 Gönder", command=gonder, font=("Segoe UI", 10, "bold"), bg="#43a047", fg="white", activebackground="#388e3c", activeforeground="white", padx=8, pady=5, relief="flat", bd=0, cursor="hand2")
        btn_gonder.pack(side="left", padx=(0,10), ipadx=4, ipady=2)
        btn_gonder.bind('<Enter>', lambda e: on_enter(btn_gonder, "#388e3c"))
        btn_gonder.bind('<Leave>', lambda e: on_leave(btn_gonder, "#43a047"))
        btn_kapat = tk.Button(btn_frame, text="❌ Kapat", command=popup.destroy, font=("Segoe UI", 10, "bold"), bg="#d32f2f", fg="white", activebackground="#b71c1c", activeforeground="white", padx=8, pady=5, relief="flat", bd=0, cursor="hand2")
        btn_kapat.pack(side="left", padx=(10,0), ipadx=4, ipady=2)
        btn_kapat.bind('<Enter>', lambda e: on_enter(btn_kapat, "#b71c1c"))
        btn_kapat.bind('<Leave>', lambda e: on_leave(btn_kapat, "#d32f2f"))
        popup.bind('<Escape>', lambda e: popup.destroy())
        popup.bind('<Return>', lambda e: gonder())

    def excel_aktar(self):
        wb = Workbook()
        ws = wb.active
        if ws is None:
            messagebox.showerror("Hata", "Excel çalışma sayfası oluşturulamadı.")
            return
        ws.title = "Kategoriler ve Ödüller"
        ws.append(["S.NO", "KATEGORİ", "ÖDÜLLER"])
        for i, (kategori_var, odul_list) in enumerate(zip(self.kategori_adlari_vars, self.odul_listeleri)):
            oduller = []
            for odul in odul_list:
                if odul['tur'] == 'Katılım Madalyası':
                    oduller.append(f"Katılım Madalyası (adet: {odul['adet']})")
                elif odul['tur'] == 'En İyi Kadın':
                    oduller.append(f"En İyi Kadın [{odul['baslangic']}-{odul['bitis']}]")
                elif odul['tur'] == 'Kupa':
                    oduller.append(f"Kupa [{odul['baslangic']}-{odul['bitis']}]")
                elif odul['tur'] == 'Madalya':
                    oduller.append(f"Madalya [{odul['baslangic']}-{odul['bitis']}]")
                elif odul['tur'] == 'Diğer':
                    if 'baslangic' in odul and 'bitis' in odul:
                        oduller.append(f"{odul['metin']} [{odul['baslangic']}-{odul['bitis']}]")
                    else:
                        oduller.append(odul['metin'])
                else:
                    oduller.append(odul['tur'])
            ws.append([str(i+1), kategori_var.get(), ", ".join(oduller)])
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyası", "*.xlsx")], title="Excel Olarak Kaydet")
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Başarılı", f"Excel dosyası kaydedildi:\n{file_path}")

    def next_step(self):
        if self.step < 3:
            self.step += 1
        self.build_step()

    def prev_step(self):
        if self.step > 0:
            self.step -= 1
        self.build_step()

    def update_btn_color(self, var, btn):
        if btn is not None:
            btn.config(bg=var.get())

    def etiketler_olustur_sirali(self):
        etiketler_madalya = []
        etiketler_diger = []
        etiketler_kupa = []
        il_adi = self.il_adi.get()
        turnuva_adi = (self.turnuva_adi.get().strip() or "") + " SATRANÇ TURNUVASI"
        bas_tarih = self.bas_tarih.get()
        bit_tarih = self.bit_tarih.get()
        tarih = bas_tarih if bas_tarih == bit_tarih else f"{bas_tarih} - {bit_tarih}"
        # None koruması
        kategori_adlari_vars = self.kategori_adlari_vars if self.kategori_adlari_vars is not None else []
        odul_listeleri = self.odul_listeleri if self.odul_listeleri is not None else []
        for i in range(self.kategori_sayisi.get()):
            if i >= len(kategori_adlari_vars) or i >= len(odul_listeleri):
                continue
            kategori = (kategori_adlari_vars[i].get().strip() or f"Kategori {i+1}") + " KATEGORİSİ"
            odul_listesi = odul_listeleri[i]
            for odul in odul_listesi:
                tur = odul["tur"]
                if tur == "Katılım Madalyası":
                    for _ in range(odul["adet"]):
                        etiketler_madalya.append({
                            "il": il_adi,
                            "turnuva_adi": turnuva_adi,
                            "kategori": kategori,
                            "derece": "Katılım Madalyası",
                            "tarih": tarih,
                            "w": float(self.madalya_w.get())*cm,
                            "h": float(self.madalya_h.get())*cm,
                            "font": self.madalya_font.get(),
                            "fontsize": self.madalya_fontsize.get(),
                            "color": self.madalya_color.get(),
                            "sol_logo": self.sol_logo_path.get(),
                            "sag_logo": self.sag_logo_path.get(),
                        })
                elif tur == "Kupa":
                    for derece in range(odul["baslangic"], odul["bitis"]+1):
                        etiketler_kupa.append({
                            "il": il_adi,
                            "turnuva_adi": turnuva_adi,
                            "kategori": kategori,
                            "derece": derece_metni(derece),
                            "tarih": tarih,
                            "w": float(self.kupa_w.get())*cm,
                            "h": float(self.kupa_h.get())*cm,
                            "font": self.kupa_font.get(),
                            "fontsize": self.kupa_fontsize.get(),
                            "color": self.kupa_color.get(),
                            "sol_logo": self.sol_logo_path.get(),
                            "sag_logo": self.sag_logo_path.get(),
                        })
                elif tur == "Madalya":
                    for derece in range(odul["baslangic"], odul["bitis"]+1):
                        etiketler_madalya.append({
                            "il": il_adi,
                            "turnuva_adi": turnuva_adi,
                            "kategori": kategori,
                            "derece": derece_metni(derece),
                            "tarih": tarih,
                            "w": float(self.madalya_w.get())*cm,
                            "h": float(self.madalya_h.get())*cm,
                            "font": self.madalya_font.get(),
                            "fontsize": self.madalya_fontsize.get(),
                            "color": self.madalya_color.get(),
                            "sol_logo": self.sol_logo_path.get(),
                            "sag_logo": self.sag_logo_path.get(),
                        })
                else:
                    # Diğer ödüller
                    metin = odul.get('metin', odul.get('tur', 'Ödül'))
                    if 'baslangic' in odul and 'bitis' in odul:
                        for derece in range(odul['baslangic'], odul['bitis']+1):
                            etiketler_diger.append({
                                "il": il_adi,
                                "turnuva_adi": turnuva_adi,
                                "kategori": kategori,
                                "derece": f"{metin} {derece_metni(derece)}",
                                "tarih": tarih,
                                "w": float(self.madalya_w.get())*cm,
                                "h": float(self.madalya_h.get())*cm,
                                "font": self.madalya_font.get(),
                                "fontsize": self.madalya_fontsize.get(),
                                "color": self.madalya_color.get(),
                                "sol_logo": self.sol_logo_path.get(),
                                "sag_logo": self.sag_logo_path.get(),
                            })
                    else:
                        etiketler_diger.append({
                            "il": il_adi,
                            "turnuva_adi": turnuva_adi,
                            "kategori": kategori,
                            "derece": metin,
                            "tarih": tarih,
                            "w": float(self.madalya_w.get())*cm,
                            "h": float(self.madalya_h.get())*cm,
                            "font": self.madalya_font.get(),
                            "fontsize": self.madalya_fontsize.get(),
                            "color": self.madalya_color.get(),
                            "sol_logo": self.sol_logo_path.get(),
                            "sag_logo": self.sag_logo_path.get(),
                        })
        return etiketler_madalya + etiketler_diger + etiketler_kupa

    def metinleri_yaz(self, c, x, y, w, h, etiket, font_name=None, font_size=None, font_color="#000000"):
        padding = 0.2 * cm
        max_width = w - 2 * padding
        max_height = h - 2 * padding
        try:
            max_font = int(font_size) if font_size is not None else 10
        except Exception:
            max_font = 10
        min_font = 2
        il = etiket["il"]
        turnuva_adi = etiket["turnuva_adi"]
        kategori = etiket["kategori"]
        derece = etiket["derece"]
        tarih = etiket["tarih"]
        best_font = min_font
        best_lines = None
        for fs in range(max_font, min_font-1, -1):
            line_spacing = 1.0  # Daha sıkı satır aralığı
            extra_spacing = 0.0
            lines = [il, turnuva_adi, kategori]
            if derece:
                lines.append(derece)
            lines.append(tarih)
            wrapped_lines = []
            for idx, line in enumerate(lines):
                if c.stringWidth(line, font_name, fs) > max_width:
                    import textwrap
                    wrapper = textwrap.TextWrapper(width=max(1, int(max_width // (fs*0.6))))
                    wrapped_lines.extend(wrapper.wrap(line))
                else:
                    wrapped_lines.append(line)
            total_height = len(wrapped_lines) * fs * line_spacing + extra_spacing
            max_line_width = max([c.stringWidth(line, font_name, fs) for line in wrapped_lines])
            if total_height <= max_height and max_line_width <= max_width:
                best_font = fs
                best_lines = wrapped_lines
                break
        if best_lines is None:
            best_lines = [il, turnuva_adi, kategori]
            if derece:
                best_lines.append(derece)
            best_lines.append(tarih)
            best_font = min_font
            line_spacing = 1.0
            extra_spacing = 0.0
            total_height = len(best_lines) * best_font * line_spacing + extra_spacing
        else:
            line_spacing = 1.0
            total_height = len(best_lines) * best_font * line_spacing + extra_spacing
        y_start = y + (h + total_height) / 2 - best_font * line_spacing / 2
        c.setFillColor(HexColor(font_color))
        for i, line in enumerate(best_lines):
            c.setFont(font_name, best_font)
            c.drawCentredString(x + w / 2, y_start - i * best_font * line_spacing, line)

    def odul_ekle_popup(self, kategori_idx):
        popup = tk.Toplevel(self.root)
        popup.title("Ödül Ekle")
        popup.configure(bg="#f5f7fa")
        popup.resizable(False, False)
        popup.geometry("340x220+%d+%d" % (self.root.winfo_rootx()+120, self.root.winfo_rooty()+120))
        popup.bind('<Escape>', lambda e: popup.destroy())
        # Modern başlık
        tk.Label(popup, text="Ödül Türü:", font=("Segoe UI", 12, "bold"), bg="#f5f7fa", fg="#1976d2").pack(pady=(16, 6))
        tur_var = tk.StringVar(value="Madalya")
        turler = ["Madalya", "Kupa", "En İyi Kadın", "Katılım Madalyası", "Diğer"]
        style = ttk.Style(popup)
        style.theme_use('clam')
        style.configure('Modern.TCombobox', font=("Segoe UI", 11), fieldbackground="#fff", background="#e3f2fd", bordercolor="#2196f3", borderwidth=1, relief="flat")
        tur_menu = ttk.Combobox(popup, textvariable=tur_var, values=turler, state="readonly", style='Modern.TCombobox', font=("Segoe UI", 11))
        tur_menu.pack(pady=4, ipadx=6, ipady=2)
        frame = tk.Frame(popup, bg="#f5f7fa")
        frame.pack(pady=8)
        bas_var = tk.IntVar(value=1)
        bit_var = tk.IntVar(value=1)
        adet_var = tk.IntVar(value=1)
        metin_var = tk.StringVar()
        def update_fields(*args):
            # Önce frame içindeki tüm widget'ları yok et
            for child in frame.winfo_children():
                child.destroy()
            if tur_var.get() == "Katılım Madalyası":
                adet_label = tk.Label(frame, text="Adet:", font=("Segoe UI", 10), bg="#f5f7fa")
                adet_spin = tk.Spinbox(frame, from_=1, to=100, width=5, textvariable=adet_var, font=("Segoe UI", 10), relief="ridge", bd=1, state='normal')
                adet_label.pack(side=tk.LEFT, padx=(0,2))
                adet_spin.pack(side=tk.LEFT, padx=(0,8))
            elif tur_var.get() == "Diğer":
                # İlk satır: Etiket Metni
                row1 = tk.Frame(frame, bg="#f5f7fa")
                metin_label = tk.Label(row1, text="Etiket Metni:", font=("Segoe UI", 10), bg="#f5f7fa")
                metin_entry = tk.Entry(row1, textvariable=metin_var, width=20, font=("Segoe UI", 10), relief="ridge", bd=1)
                metin_label.pack(side=tk.LEFT, padx=(0,2))
                metin_entry.pack(side=tk.LEFT, padx=(0,8))
                row1.pack(anchor="w", pady=(0,2))
                # İkinci satır: Başlangıç ve Bitiş
                row2 = tk.Frame(frame, bg="#f5f7fa")
                bas_label = tk.Label(row2, text="Başlangıç:", font=("Segoe UI", 10), bg="#f5f7fa")
                bas_spin = tk.Spinbox(row2, from_=1, to=100, width=5, textvariable=bas_var, font=("Segoe UI", 10), relief="ridge", bd=1, state='normal')
                bit_label = tk.Label(row2, text="Bitiş:", font=("Segoe UI", 10), bg="#f5f7fa")
                bit_spin = tk.Spinbox(row2, from_=1, to=100, width=5, textvariable=bit_var, font=("Segoe UI", 10), relief="ridge", bd=1, state='normal')
                bas_label.pack(side=tk.LEFT, padx=(0,2))
                bas_spin.pack(side=tk.LEFT, padx=(0,8))
                bit_label.pack(side=tk.LEFT, padx=(0,2))
                bit_spin.pack(side=tk.LEFT, padx=(0,8))
                row2.pack(anchor="w")
            else:
                bas_label = tk.Label(frame, text="Başlangıç:", font=("Segoe UI", 10), bg="#f5f7fa")
                bas_spin = tk.Spinbox(frame, from_=1, to=100, width=5, textvariable=bas_var, font=("Segoe UI", 10), relief="ridge", bd=1, state='normal')
                bit_label = tk.Label(frame, text="Bitiş:", font=("Segoe UI", 10), bg="#f5f7fa")
                bit_spin = tk.Spinbox(frame, from_=1, to=100, width=5, textvariable=bit_var, font=("Segoe UI", 10), relief="ridge", bd=1, state='normal')
                bas_label.pack(side=tk.LEFT, padx=(0,2))
                bas_spin.pack(side=tk.LEFT, padx=(0,8))
                bit_label.pack(side=tk.LEFT, padx=(0,2))
                bit_spin.pack(side=tk.LEFT, padx=(0,8))
        tur_var.trace_add('write', update_fields)
        tur_menu.bind('<<ComboboxSelected>>', lambda e: update_fields())
        update_fields()
        btn_frame = tk.Frame(popup, bg="#f5f7fa")
        btn_frame.pack(pady=16)
        def on_enter(e, btn): btn.config(bg="#43a047")
        def on_leave(e, btn): btn.config(bg="#2196f3")
        ekle_btn = tk.Button(btn_frame, text="Ekle", font=("Segoe UI", 11, "bold"), bg="#2196f3", fg="white", activebackground="#43a047", activeforeground="white", relief="flat", bd=0, padx=18, pady=4, cursor="hand2", command=lambda: ekle())
        ekle_btn.pack(side=tk.LEFT, padx=8)
        ekle_btn.bind('<Enter>', lambda e: on_enter(e, ekle_btn))
        ekle_btn.bind('<Leave>', lambda e: on_leave(e, ekle_btn))
        iptal_btn = tk.Button(btn_frame, text="İptal", font=("Segoe UI", 11, "bold"), bg="#bdbdbd", fg="white", activebackground="#757575", activeforeground="white", relief="flat", bd=0, padx=18, pady=4, cursor="hand2", command=popup.destroy)
        iptal_btn.pack(side=tk.LEFT, padx=8)
        iptal_btn.bind('<Enter>', lambda e: iptal_btn.config(bg="#757575"))
        iptal_btn.bind('<Leave>', lambda e: iptal_btn.config(bg="#bdbdbd"))
        def ekle():
            if tur_var.get() == "Katılım Madalyası":
                adet = adet_var.get()
                if adet < 1:
                    self.show_error("Hata", "Adet en az 1 olmalı.")
                    return
                self.odul_listeleri[kategori_idx].append({"tur": "Katılım Madalyası", "adet": adet})
            elif tur_var.get() == "Diğer":
                metin = metin_var.get().strip()
                bas = bas_var.get()
                bit = bit_var.get()
                if not metin:
                    self.show_error("Hata", "Etiket metni giriniz.")
                    return
                if bit >= bas:
                    self.odul_listeleri[kategori_idx].append({
                        "tur": "Diğer",
                        "metin": metin,
                        "baslangic": bas,
                        "bitis": bit
                    })
                else:
                    self.odul_listeleri[kategori_idx].append({"tur": "Diğer", "metin": metin})
            else:
                bas = bas_var.get()
                bit = bit_var.get()
                if bit < bas:
                    self.show_error("Hata", "Bitiş, başlangıçtan küçük olamaz.")
                    return
                self.odul_listeleri[kategori_idx].append({"tur": tur_var.get(), "baslangic": bas, "bitis": bit})
            popup.destroy()
            self.build_step()
        popup.bind('<Escape>', lambda e: popup.destroy())
        popup.bind('<Return>', lambda e: ekle())

    def odul_kaldir(self, kategori_idx, odul_idx):
        del self.odul_listeleri[kategori_idx][odul_idx]
        self.build_step()

    def odul_duzenle_popup(self, kategori_idx, odul_idx):
        odul = self.odul_listeleri[kategori_idx][odul_idx]
        popup = tk.Toplevel(self.root)
        popup.title("Ödül Düzenle")
        popup.configure(bg="#f5f7fa")
        popup.resizable(False, False)
        popup.geometry("340x220+%d+%d" % (self.root.winfo_rootx()+120, self.root.winfo_rooty()+120))
        popup.bind('<Escape>', lambda e: popup.destroy())
        tk.Label(popup, text="Ödül Türü:", font=("Segoe UI", 12, "bold"), bg="#f5f7fa", fg="#1976d2").pack(pady=(16, 6))
        tur_var = tk.StringVar(value=odul.get("tur", "Madalya"))
        turler = ["Madalya", "Kupa", "En İyi Kadın", "Katılım Madalyası", "Diğer"]
        style = ttk.Style(popup)
        style.theme_use('clam')
        style.configure('Modern.TCombobox', font=("Segoe UI", 11), fieldbackground="#fff", background="#e3f2fd", bordercolor="#2196f3", borderwidth=1, relief="flat")
        tur_menu = ttk.Combobox(popup, textvariable=tur_var, values=turler, state="readonly", style='Modern.TCombobox', font=("Segoe UI", 11))
        tur_menu.pack(pady=4, ipadx=6, ipady=2)
        frame = tk.Frame(popup, bg="#f5f7fa")
        frame.pack(pady=8)
        bas_var = tk.IntVar(value=odul.get("baslangic", 1))
        bit_var = tk.IntVar(value=odul.get("bitis", 1))
        adet_var = tk.IntVar(value=odul.get("adet", 1))
        metin_var = tk.StringVar(value=odul.get("metin", ""))
        def update_fields(*args):
            for child in frame.winfo_children():
                child.destroy()
            if tur_var.get() == "Katılım Madalyası":
                adet_label = tk.Label(frame, text="Adet:", font=("Segoe UI", 10), bg="#f5f7fa")
                adet_spin = tk.Spinbox(frame, from_=1, to=100, width=5, textvariable=adet_var, font=("Segoe UI", 10), relief="ridge", bd=1, state='normal')
                adet_label.pack(side=tk.LEFT, padx=(0,2))
                adet_spin.pack(side=tk.LEFT, padx=(0,8))
            elif tur_var.get() == "Diğer":
                row1 = tk.Frame(frame, bg="#f5f7fa")
                metin_label = tk.Label(row1, text="Etiket Metni:", font=("Segoe UI", 10), bg="#f5f7fa")
                metin_entry = tk.Entry(row1, textvariable=metin_var, width=20, font=("Segoe UI", 10), relief="ridge", bd=1)
                metin_label.pack(side=tk.LEFT, padx=(0,2))
                metin_entry.pack(side=tk.LEFT, padx=(0,8))
                row1.pack(anchor="w", pady=(0,2))
                row2 = tk.Frame(frame, bg="#f5f7fa")
                bas_label = tk.Label(row2, text="Başlangıç:", font=("Segoe UI", 10), bg="#f5f7fa")
                bas_spin = tk.Spinbox(row2, from_=1, to=100, width=5, textvariable=bas_var, font=("Segoe UI", 10), relief="ridge", bd=1, state='normal')
                bit_label = tk.Label(row2, text="Bitiş:", font=("Segoe UI", 10), bg="#f5f7fa")
                bit_spin = tk.Spinbox(row2, from_=1, to=100, width=5, textvariable=bit_var, font=("Segoe UI", 10), relief="ridge", bd=1, state='normal')
                bas_label.pack(side=tk.LEFT, padx=(0,2))
                bas_spin.pack(side=tk.LEFT, padx=(0,8))
                bit_label.pack(side=tk.LEFT, padx=(0,2))
                bit_spin.pack(side=tk.LEFT, padx=(0,8))
                row2.pack(anchor="w")
            else:
                bas_label = tk.Label(frame, text="Başlangıç:", font=("Segoe UI", 10), bg="#f5f7fa")
                bas_spin = tk.Spinbox(frame, from_=1, to=100, width=5, textvariable=bas_var, font=("Segoe UI", 10), relief="ridge", bd=1, state='normal')
                bit_label = tk.Label(frame, text="Bitiş:", font=("Segoe UI", 10), bg="#f5f7fa")
                bit_spin = tk.Spinbox(frame, from_=1, to=100, width=5, textvariable=bit_var, font=("Segoe UI", 10), relief="ridge", bd=1, state='normal')
                bas_label.pack(side=tk.LEFT, padx=(0,2))
                bas_spin.pack(side=tk.LEFT, padx=(0,8))
                bit_label.pack(side=tk.LEFT, padx=(0,2))
                bit_spin.pack(side=tk.LEFT, padx=(0,8))
        tur_var.trace_add('write', update_fields)
        tur_menu.bind('<<ComboboxSelected>>', lambda e: update_fields())
        update_fields()
        btn_frame = tk.Frame(popup, bg="#f5f7fa")
        btn_frame.pack(pady=16)
        def on_enter(e, btn): btn.config(bg="#43a047")
        def on_leave(e, btn): btn.config(bg="#2196f3")
        kaydet_btn = tk.Button(btn_frame, text="Kaydet", font=("Segoe UI", 11, "bold"), bg="#2196f3", fg="white", activebackground="#43a047", activeforeground="white", relief="flat", bd=0, padx=18, pady=4, cursor="hand2", command=lambda: kaydet())
        kaydet_btn.pack(side=tk.LEFT, padx=8)
        kaydet_btn.bind('<Enter>', lambda e: on_enter(e, kaydet_btn))
        kaydet_btn.bind('<Leave>', lambda e: on_leave(e, kaydet_btn))
        iptal_btn = tk.Button(btn_frame, text="İptal", font=("Segoe UI", 11, "bold"), bg="#bdbdbd", fg="white", activebackground="#757575", activeforeground="white", relief="flat", bd=0, padx=18, pady=4, cursor="hand2", command=popup.destroy)
        iptal_btn.pack(side=tk.LEFT, padx=8)
        iptal_btn.bind('<Enter>', lambda e: iptal_btn.config(bg="#757575"))
        iptal_btn.bind('<Leave>', lambda e: iptal_btn.config(bg="#bdbdbd"))
        def kaydet():
            if tur_var.get() == "Katılım Madalyası":
                adet = adet_var.get()
                if adet < 1:
                    self.show_error("Hata", "Adet en az 1 olmalı.")
                    return
                self.odul_listeleri[kategori_idx][odul_idx] = {"tur": "Katılım Madalyası", "adet": adet}
            elif tur_var.get() == "Diğer":
                metin = metin_var.get().strip()
                bas = bas_var.get()
                bit = bit_var.get()
                if not metin:
                    self.show_error("Hata", "Etiket metni giriniz.")
                    return
                if bit >= bas:
                    self.odul_listeleri[kategori_idx][odul_idx] = {
                        "tur": "Diğer",
                        "metin": metin,
                        "baslangic": bas,
                        "bitis": bit
                    }
                else:
                    self.odul_listeleri[kategori_idx][odul_idx] = {"tur": "Diğer", "metin": metin}
            else:
                bas = bas_var.get()
                bit = bit_var.get()
                if bit < bas:
                    self.show_error("Hata", "Bitiş, başlangıçtan küçük olamaz.")
                    return
                self.odul_listeleri[kategori_idx][odul_idx] = {"tur": tur_var.get(), "baslangic": bas, "bitis": bit}
            popup.destroy()
            self.build_step()
        popup.bind('<Escape>', lambda e: popup.destroy())
        popup.bind('<Return>', lambda e: kaydet())

    def pdf_onizle(self):
        etiketler = self.etiketler_olustur_sirali()
        if not etiketler:
            self.show_info("Uyarı", "Önizlenecek bir etiket yok.")
            return
        temp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        self.pdf_yaz(etiketler, temp.name)
        self.preview_tempfile = temp.name
        try:
            if os.name == 'nt':
                os.startfile(temp.name)
            else:
                subprocess.Popen(['xdg-open', temp.name])
        except Exception as e:
            self.show_error("Hata", f"PDF açılamadı: {e}")

    def pdf_olustur(self):
        etiketler = self.etiketler_olustur_sirali()
        if not etiketler:
            self.show_info("Uyarı", "Oluşturulacak bir etiket yok.")
            return
        pdf_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Dosyası", "*.pdf")])
        if not pdf_path:
            return
        self.pdf_yaz(etiketler, pdf_path)
        self.pdf_path = pdf_path
        self.show_info("Başarılı", f"PDF başarıyla kaydedildi:\n{pdf_path}")
        try:
            if os.name == 'nt':
                os.startfile(pdf_path)
            else:
                subprocess.Popen(['xdg-open', pdf_path])
        except Exception as e:
            self.show_error("Hata", f"PDF açılamadı: {e}")

    def pdf_yaz(self, etiketler, pdf_path):
        c = canvas.Canvas(pdf_path, pagesize=A4)
        page_width, page_height = A4
        left_margin = 1 * cm
        right_margin = 1 * cm
        etiket_aralik_x = 0.5 * cm
        etiket_aralik_y = 0.5 * cm
        self.etiketleri_yaz(etiketler, c, page_width, page_height, left_margin, right_margin, etiket_aralik_x, etiket_aralik_y)
        c.save()

    def etiketleri_yaz(self, etiketler, c, page_width, page_height, left_margin, right_margin, etiket_aralik_x, etiket_aralik_y):
        etiket_idx = 0
        total = len(etiketler)
        while etiket_idx < total:
            y_cursor = page_height - left_margin
            while y_cursor > left_margin and etiket_idx < total:
                x_cursor = left_margin
                max_row_height = 0
                # Madalya etiketleri için 5'li satır zorunluluğu
                if etiketler[etiket_idx]["derece"] == "Katılım Madalyası":
                    etiketler_bu_satir = []
                    kalan = total - etiket_idx
                    for i in range(min(5, kalan)):
                        etiket = etiketler[etiket_idx + i]
                        if etiket["derece"] != "Katılım Madalyası":
                            break
                        w = etiket["w"]
                        if x_cursor + w > page_width - right_margin:
                            break
                        etiketler_bu_satir.append(etiket)
                        x_cursor += w + etiket_aralik_x
                    if len(etiketler_bu_satir) == 0:
                        break
                    x_cursor = left_margin
                    for etiket in etiketler_bu_satir:
                        w = etiket["w"]
                        h = etiket["h"]
                        font = etiket.get("font", "DejaVuSans")
                        fontsize = etiket.get("fontsize", "10")
                        color = etiket.get("color", "#000000")
                        c.setLineWidth(1)
                        c.rect(x_cursor, y_cursor - h, w, h)
                        sol_logo = etiket.get("sol_logo", "")
                        sag_logo = etiket.get("sag_logo", "")
                        logo_h = min(1.2*cm, h*0.25)
                        logo_w = logo_h
                        logo_y = y_cursor - 2 - logo_h
                        if sol_logo and os.path.exists(sol_logo):
                            try:
                                img = Image.open(sol_logo)
                                img = img.resize((int(logo_w), int(logo_h)), Image.Resampling.LANCZOS)
                                c.drawInlineImage(img, x_cursor+2, logo_y, width=logo_w, height=logo_h)
                            except Exception as e:
                                pass
                        if sag_logo and os.path.exists(sag_logo):
                            try:
                                img = Image.open(sag_logo)
                                img = img.resize((int(logo_w), int(logo_h)), Image.Resampling.LANCZOS)
                                c.drawInlineImage(img, x_cursor+w-logo_w-2, logo_y, width=logo_w, height=logo_h)
                            except Exception as e:
                                pass
                        self.metinleri_yaz(c, x_cursor, y_cursor - h, w, h, etiket, font, fontsize, color)
                        max_row_height = max(max_row_height, h)
                        x_cursor += w + etiket_aralik_x
                        etiket_idx += 1
                    y_cursor -= max_row_height + etiket_aralik_y
                else:
                    # Diğer ödüller ve kupalar için mevcut mantık
                    while x_cursor < page_width - right_margin and etiket_idx < total:
                        w = etiketler[etiket_idx]["w"]
                        h = etiketler[etiket_idx]["h"]
                        font = etiketler[etiket_idx].get("font", "DejaVuSans")
                        fontsize = etiketler[etiket_idx].get("fontsize", "10")
                        color = etiketler[etiket_idx].get("color", "#000000")
                        if x_cursor + w > page_width - right_margin or y_cursor - h < left_margin:
                            break
                        c.setLineWidth(1)
                        c.rect(x_cursor, y_cursor - h, w, h)
                        sol_logo = etiketler[etiket_idx].get("sol_logo", "")
                        sag_logo = etiketler[etiket_idx].get("sag_logo", "")
                        logo_h = min(1.2*cm, h*0.25)
                        logo_w = logo_h
                        logo_y = y_cursor - 2 - logo_h
                        if sol_logo and os.path.exists(sol_logo):
                            try:
                                img = Image.open(sol_logo)
                                img = img.resize((int(logo_w), int(logo_h)), Image.Resampling.LANCZOS)
                                c.drawInlineImage(img, x_cursor+2, logo_y, width=logo_w, height=logo_h)
                            except Exception as e:
                                pass
                        if sag_logo and os.path.exists(sag_logo):
                            try:
                                img = Image.open(sag_logo)
                                img = img.resize((int(logo_w), int(logo_h)), Image.Resampling.LANCZOS)
                                c.drawInlineImage(img, x_cursor+w-logo_w-2, logo_y, width=logo_w, height=logo_h)
                            except Exception as e:
                                pass
                        self.metinleri_yaz(c, x_cursor, y_cursor - h, w, h, etiketler[etiket_idx], font, fontsize, color)
                        max_row_height = max(max_row_height, h)
                        x_cursor += w + etiket_aralik_x
                        etiket_idx += 1
                    y_cursor -= max_row_height + etiket_aralik_y
            if etiket_idx < total:
                c.showPage()

    def pick_color(self, var, btn):
        from tkinter import colorchooser
        renk = colorchooser.askcolor(title="Renk Seç", initialcolor=var.get())
        if renk[1]:
            var.set(renk[1])

    def font_popup(self, var, btn, parent):
        popup = tk.Toplevel(parent)
        popup.title("Yazı Tipi Seç")
        popup.transient(parent)
        popup.grab_set()
        popup.geometry("340x440")
        popup.bind('<Escape>', lambda e: popup.destroy())
        sample_text = tk.Label(popup, text="Örnek Yazı: ABC abc 123", font=(var.get(), 18, "bold"), pady=12)
        sample_text.pack(pady=(10, 0))
        listbox = tk.Listbox(popup, font=("Arial", 12), height=15)
        for f in self.font_list:
            listbox.insert(tk.END, f)
        listbox.pack(expand=True, fill="both", padx=10, pady=10)
        def update_sample(event=None):
            sel = listbox.curselection()
            if sel:
                font_name = listbox.get(sel[0])
                sample_text.config(font=(font_name, 18, "bold"))
        listbox.bind("<<ListboxSelect>>", update_sample)
        def select_font(event=None):
            sel = listbox.curselection()
            if sel:
                font_name = listbox.get(sel[0])
                var.set(font_name)
                # Ana ekrandaki butonun textini de güncelle
                if btn is not None:
                    btn.config(text=font_name, font=(font_name, 11))
                popup.destroy()
        listbox.bind("<Double-Button-1>", select_font)
        tk.Button(popup, text="Seç", command=select_font, font=("Arial", 11, "bold"), bg="#1976d2", fg="white").pack(pady=8)
        # İlk açılışta seçili fontu örnekte göster
        try:
            idx = self.font_list.index(var.get())
            listbox.selection_set(idx)
            listbox.see(idx)
            sample_text.config(font=(var.get(), 18, "bold"))
        except Exception:
            pass

    def show_info(self, title, message, info_type="info"):
        color = {"info": "#2196f3", "warning": "#ff9800", "error": "#e53935"}.get(info_type, "#2196f3")
        icon_char = {"info": "ℹ️", "warning": "⚠️", "error": "❌"}.get(info_type, "ℹ️")
        top = tk.Toplevel(self.root)
        top.title(title)
        top.configure(bg="#f8fafc", bd=0, highlightthickness=0)
        top.geometry("380x170")
        top.resizable(False, False)
        # Çerçeve efekti
        frame = tk.Frame(top, bg="#fff", bd=0, highlightbackground="#b0bec5", highlightthickness=2)
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        # Üstte ikon ve başlık
        header = tk.Frame(frame, bg="#fff")
        header.pack(fill="x", pady=(10, 0), padx=10)
        icon = tk.Label(header, text=icon_char, font=(APP_FONT_FAMILY, 20, "bold"), fg=color, bg="#fff")
        icon.pack(side="left", padx=(0, 8))
        lbl_title = tk.Label(header, text=title, font=(APP_FONT_FAMILY, 13, "bold"), fg=color, bg="#fff")
        lbl_title.pack(side="left", anchor="w")
        # Mesaj
        msg = tk.Label(frame, text=message, font=(APP_FONT_FAMILY, 11), bg="#fff", fg="#333", wraplength=320, justify="center")
        msg.pack(pady=(8, 12), padx=10)
        # Buton
        btn = tk.Button(
            frame, text="Kapat", command=top.destroy,
            font=(APP_FONT_FAMILY, 13, "bold"), relief="flat",
            bg="#d32f2f", fg="white", activebackground="#b71c1c", activeforeground="white",
            bd=0, padx=24, pady=8
        )
        btn.pack(pady=(0, 8))
        top.bind('<Escape>', lambda e: top.destroy())
        top.transient(self.root)
        btn.focus_set()

    def show_error(self, title, message):
        messagebox.showerror(title, message)
        logging.error(f"{title}: {message}")

    def sayfa_degistir(self, delta):
        self.kategori_sayfa += delta
        if self.kategori_sayfa < 0:
            self.kategori_sayfa = 0
        self.build_step()

    def styled_button(self, parent, nav=False, **kwargs):
        kwargs.setdefault("font", BUTTON_FONT)
        if nav:
            kwargs.setdefault("bg", SECONDARY_COLOR)
            kwargs.setdefault("activebackground", PRIMARY_COLOR)
        else:
            kwargs.setdefault("bg", PRIMARY_COLOR)
            kwargs.setdefault("activebackground", SECONDARY_COLOR)
        kwargs.setdefault("fg", "white")
        kwargs.setdefault("activeforeground", "white")
        kwargs.setdefault("bd", 0)
        kwargs.setdefault("relief", "ridge")
        kwargs.setdefault("padx", 10)
        kwargs.setdefault("pady", 5)
        kwargs.setdefault("highlightthickness", 0)
        btn = tk.Button(parent, **kwargs)
        btn.configure(cursor="hand2", bd=0, relief="flat", highlightbackground=kwargs["bg"])
        btn.bind('<Enter>', lambda e: btn.config(bg=kwargs["activebackground"]))
        btn.bind('<Leave>', lambda e: btn.config(bg=kwargs["bg"]))
        btn.config(borderwidth=0)
        return btn

    def toplu_katilim_popup(self):
        popup = tk.Toplevel(self.root)
        popup.title("Toplu Katılım Belgesi Oluştur")
        popup.configure(bg="#f7f7f7")
        popup.geometry("520x560")
        popup.attributes("-topmost", True)
        popup.after(100, lambda: popup.attributes("-topmost", False))
        popup.bind('<Escape>', lambda e: popup.destroy())
        # Kart benzeri ana frame
        card = tk.Frame(popup, bg="#fff", bd=2, relief="ridge", highlightbackground="#e0e0e0", highlightthickness=1)
        card.pack(padx=24, pady=24, fill="both", expand=True)
        # Başlık
        tk.Label(card, text="Toplu Katılım Belgesi Oluştur", font=("Segoe UI", 15, "bold"), bg="#fff").pack(pady=(18, 6))
        tk.Frame(card, bg="#e0e0e0", height=2).pack(fill="x", padx=8, pady=(0, 12))
        form = tk.Frame(card, bg="#fff")
        form.pack(padx=12, pady=6, fill="x")
        # --- Önce yardımcı fonksiyonlar ve değişkenler tanımlanır ---
        self.toplu_isim_kategori_listesi = []
        self.toplu_dosya_adi = tk.StringVar(value="Henüz dosya seçilmedi")
        self.toplu_kisi_sayisi = tk.StringVar(value="Kişi sayısı: 0")
        logo_sol_var = tk.StringVar()
        logo_sag_var = tk.StringVar()
        def sablon_indir():
            from tkinter import filedialog
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            if ws is None:
                messagebox.showerror("Hata", "Excel çalışma sayfası oluşturulamadı.")
                return
            ws.title = "Katılım Listesi"
            ws.append(["İsim", "Kategori"])
            ws.append(["Ali Yılmaz", "Küçükler"])
            ws.append(["Ayşe Demir", "Yıldızlar"])
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyası", "*.xlsx")], title="Boş Şablon Kaydet")
            if file_path:
                wb.save(file_path)
        def dosya_sec():
            from tkinter import filedialog
            popup.attributes("-topmost", False)  # Dosya seçmeden önce normal moda al
            file_path = filedialog.askopenfilename(filetypes=[("Excel Dosyası", "*.xlsx")])
            popup.attributes("-topmost", True)   # Dosya seçtikten sonra tekrar en üste al
            if not file_path:
                return
            isim_kategori = []
            wb = load_workbook(file_path)
            ws = wb.active
            if ws is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1]:
                        isim_kategori.append((str(row[0]), str(row[1])))
            self.toplu_isim_kategori_listesi = isim_kategori
            self.toplu_dosya_adi.set(os.path.basename(file_path))
            self.toplu_kisi_sayisi.set(f"Kişi sayısı: {len(isim_kategori)}")
        def tarih_sec(entry):
            def tarih_aralik_secici(parent, entry):
                top = tk.Toplevel(parent)
                top.title("Tarih Aralığı Seç")
                cal1 = Calendar(top, selectmode='day', locale='tr_TR')
                cal1.pack(padx=10, pady=5)
                tk.Label(top, text="Başlangıç Tarihi").pack()
                cal2 = Calendar(top, selectmode='day', locale='tr_TR')
                cal2.pack(padx=10, pady=5)
                tk.Label(top, text="Bitiş Tarihi").pack()
                def sec():
                    t1 = cal1.selection_get()
                    t2 = cal2.selection_get()
                    if t1 and t2:
                        t1s = t1.strftime("%d.%m.%Y")
                        t2s = t2.strftime("%d.%m.%Y")
                        if t1s == t2s:
                            entry.delete(0, tk.END)
                            entry.insert(0, t1s)
                        else:
                            entry.delete(0, tk.END)
                            entry.insert(0, f"{t1s} - {t2s}")
                        top.destroy()
                tk.Button(top, text="Seç", command=sec).pack(pady=5)
                top.bind('<Escape>', lambda e: top.destroy())
            tarih_aralik_secici(popup, entry)
        def logo_sol_sec():
            from tkinter import filedialog
            path = filedialog.askopenfilename(filetypes=[("Resim Dosyası", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
            if path:
                dosya_adi = os.path.basename(path)
                hedef_yol = resource_path(dosya_adi)
                if not os.path.exists(hedef_yol):
                    try:
                        shutil.copy(path, hedef_yol)
                    except Exception:
                        pass
                logo_sol_var.set(dosya_adi)
                self.last_logo_sol = dosya_adi
        def logo_sol_kaldir():
            logo_sol_var.set("")
            self.last_logo_sol = ""
        def logo_sag_sec():
            from tkinter import filedialog
            path = filedialog.askopenfilename(filetypes=[("Resim Dosyası", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
            if path:
                dosya_adi = os.path.basename(path)
                hedef_yol = resource_path(dosya_adi)
                if not os.path.exists(hedef_yol):
                    try:
                        shutil.copy(path, hedef_yol)
                    except Exception:
                        pass
                logo_sag_var.set(dosya_adi)
                self.last_logo_sag = dosya_adi
        def logo_sag_kaldir():
            logo_sag_var.set("")
            self.last_logo_sag = ""
        def belge_olustur():
            if not self.toplu_isim_kategori_listesi:
                messagebox.showwarning("Uyarı", "Lütfen önce isim/kategori listesi dosyası seçin.")
                return
            il = entry_il.get().strip()
            turnuva = entry_turnuva.get().strip()
            tarih = entry_tarih.get().strip()
            mesaj = text_mesaj.get("1.0", "end").strip()
            logo_sol = logo_sol_var.get().strip()
            logo_sag = logo_sag_var.get().strip()
            direktor = entry_direktor.get().strip()
            bashakem = entry_bashakem.get().strip()
            from tkinter import filedialog
            pdf_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Dosyası", "*.pdf")], title="Toplu Katılım Belgeleri PDF Olarak Kaydet")
            if not pdf_path:
                return
            self.last_il = il
            self.last_turnuva = turnuva
            self.last_tarih = tarih
            self.last_mesaj = mesaj
            self.last_logo_sol = logo_sol
            self.last_logo_sag = logo_sag
            self.last_direktor = direktor
            self.last_bashakem = bashakem
            katilim_belgesi_pdf(pdf_path, self.toplu_isim_kategori_listesi, il, turnuva, tarih, logo_sol, logo_sag, direktor, bashakem, mesaj)
            messagebox.showinfo("Başarılı", f"Toplu katılım belgeleri PDF olarak kaydedildi:\n{pdf_path}")
            popup.destroy()
        # --- Sonra form ve butonlar tanımlanır ---
        # Küçük fontlar ve az padding ile form
        small_font = ("Segoe UI", 9)
        small_font_bold = ("Segoe UI", 9, "bold")
        # Boş Şablon İndir butonu
        btn_sablon = tk.Button(form, text="Boş Şablon İndir", command=sablon_indir, bg="#1976d2", fg="white", font=small_font_bold, width=14)
        btn_sablon.grid(row=0, column=0, sticky="w", pady=(0,4), columnspan=2)
        # Dosya seç
        tk.Label(form, text="İsim/Kategori Listesi (Excel):", bg="#fff", font=small_font).grid(row=1, column=0, sticky="e")
        btn_yukle = tk.Button(form, text="Dosya Seç", command=dosya_sec, font=small_font, width=8)
        btn_yukle.grid(row=1, column=1, sticky="w")
        tk.Label(form, textvariable=self.toplu_dosya_adi, fg="#888", bg="#fff", font=("Segoe UI", 8, "italic")).grid(row=2, column=0, sticky="w", columnspan=2)
        tk.Label(form, textvariable=self.toplu_kisi_sayisi, fg="#388e3c", bg="#fff", font=("Segoe UI", 8)).grid(row=3, column=0, sticky="w", columnspan=2)
        # Turnuva bilgileri ve mesaj alanı
        tk.Label(form, text="Turnuva İli:", bg="#fff", font=small_font).grid(row=4, column=0, sticky="e", pady=(8,0))
        entry_il = tk.Entry(form, width=22, font=small_font)
        entry_il.grid(row=4, column=1, sticky="w", pady=(8,0))
        tk.Label(form, text="Turnuva Adı:", bg="#fff", font=small_font).grid(row=5, column=0, sticky="e")
        entry_turnuva = tk.Entry(form, width=22, font=small_font)
        entry_turnuva.grid(row=5, column=1, sticky="w")
        tk.Label(form, text="Tarih(ler):", bg="#fff", font=small_font).grid(row=6, column=0, sticky="e")
        entry_tarih = tk.Entry(form, width=16, font=small_font)
        entry_tarih.grid(row=6, column=1, sticky="w")
        btn_tarih = tk.Button(form, text="Tarih Seç", command=lambda: tarih_sec(entry_tarih), font=small_font, width=8)
        btn_tarih.grid(row=6, column=2, sticky="w")
        tk.Label(form, text="Teşekkür/Mesaj:", bg="#fff", font=small_font).grid(row=7, column=0, sticky="ne", pady=(6,0))
        text_mesaj = tk.Text(form, width=30, height=2, font=small_font)
        text_mesaj.grid(row=7, column=1, sticky="w", pady=(6,0), columnspan=2)
        tk.Label(form, text="Örnek: Satranç sporuna verdiğiniz katkı ve sportmenliğiniz, bu etkinliğin başarılı geçmesinde önemli bir rol oynamıştır. Başarılarınızın devamını dileriz.", font=("Segoe UI", 8, "italic"), fg="#aaa", bg="#fff", wraplength=220, justify="left").grid(row=8, column=1, sticky="w", padx=(0,0), pady=(0,4), columnspan=2)
        # Logolar ve imza alanları
        tk.Label(form, text="Sol Logo:", bg="#fff", font=small_font).grid(row=9, column=0, sticky="e", pady=(8,0))
        entry_logo_sol = tk.Entry(form, textvariable=logo_sol_var, width=22, font=small_font)
        entry_logo_sol.grid(row=9, column=1, sticky="w", pady=(8,0))
        btn_logo_sol = tk.Button(form, text="Dosya Seç", command=logo_sol_sec, font=small_font, width=8)
        btn_logo_sol.grid(row=9, column=2, sticky="w", pady=(8,0))
        btn_logo_sol_kaldir = tk.Button(form, text="Kaldır", command=logo_sol_kaldir, font=small_font, width=6)
        btn_logo_sol_kaldir.grid(row=9, column=3, sticky="w", pady=(8,0))
        tk.Label(form, text="Sağ Logo:", bg="#fff", font=small_font).grid(row=10, column=0, sticky="e")
        entry_logo_sag = tk.Entry(form, textvariable=logo_sag_var, width=22, font=small_font)
        entry_logo_sag.grid(row=10, column=1, sticky="w")
        btn_logo_sag = tk.Button(form, text="Dosya Seç", command=logo_sag_sec, font=small_font, width=8)
        btn_logo_sag.grid(row=10, column=2, sticky="w")
        btn_logo_sag_kaldir = tk.Button(form, text="Kaldır", command=logo_sag_kaldir, font=small_font, width=6)
        btn_logo_sag_kaldir.grid(row=10, column=3, sticky="w")
        tk.Label(form, text="Turnuva Direktörü:", bg="#fff", font=small_font).grid(row=11, column=0, sticky="e", pady=(8,0))
        entry_direktor = tk.Entry(form, width=22, font=small_font)
        entry_direktor.grid(row=11, column=1, sticky="w", pady=(8,0), columnspan=2)
        tk.Label(form, text="Turnuva Başhakemi:", bg="#fff", font=small_font).grid(row=12, column=0, sticky="e")
        entry_bashakem = tk.Entry(form, width=22, font=small_font)
        entry_bashakem.grid(row=12, column=1, sticky="w", columnspan=2)
        btn_olustur = tk.Button(card, text="Belge Oluştur", command=belge_olustur, bg="#388e3c", fg="white", font=small_font_bold, width=14, relief="flat", bd=0, activebackground="#2e7031")
        btn_olustur.pack(pady=(10, 8))

    def show_about(self):
        top = tk.Toplevel(self.root)
        top.title("Hakkında")
        top.configure(bg=BG_COLOR)
        top.geometry("480x320")
        top.minsize(380, 220)
        top.resizable(True, True)
        top.bind('<Escape>', lambda e: top.destroy())
        icon = tk.Label(top, text="♟️", font=("Arial", 38), bg=BG_COLOR)
        icon.pack(pady=(18, 0), anchor="w")
        content_frame = tk.Frame(top, bg=BG_COLOR)
        content_frame.pack(fill="both", expand=True, padx=18, pady=6)
        msg = tk.Label(content_frame, text=(
            "Satranç Turnuvası Etiket ve Katılım Belgesi Oluşturucu\n"
            "Sürüm: 22.07.24.02\n\n"
            "Bu uygulama, satranç turnuvalarında kupa, madalya ve katılım belgesi etiketlerini hızlı ve pratik şekilde hazırlamak için geliştirilmiştir.\n\n"
            "Geliştirici: Ertuğrul Kamil ŞAHİN\n"
            "Satranç Hakemi\n\n"
            "E-posta: ertugrul.yazilim@gmail.com\n\n"
            "© 2025 Ertuğrul Kamil ŞAHİN. Tüm hakları saklıdır.\n"
        ), font=("Arial", 10), bg=BG_COLOR, justify="left", anchor="w")
        msg.pack(fill="both", expand=True, anchor="w")
        def update_wrap(event):
            wrap = max(200, content_frame.winfo_width() - 24)
            msg.config(wraplength=wrap)
        content_frame.bind("<Configure>", update_wrap)
        btn = self.styled_button(top, text="Kapat", command=top.destroy)
        btn.pack(pady=8)
        top.transient(self.root)
        top.grab_set()

    def show_help(self):
        try:
            with open("KULLANIM_KILAVUZU.txt", "r", encoding="utf-8") as f:
                help_text = f.read()
        except Exception:
            help_text = "Kullanım kılavuzu dosyası bulunamadı."
        self.show_info("Kullanım Kılavuzu", help_text)

    def show_update_info(self):
        self.show_info("Güncelleme Bilgisi", "Güncellemeleri https://github.com/kullaniciadi/guncelleme-sunucum adresinden veya geliştiriciyle iletişime geçerek takip edebilirsiniz.\n\nMevcut sürüm: " + APP_VERSION)

    def guncelleme_kontrol_et(self):
        GUNCELLEME_URL = "https://raw.githubusercontent.com/ertugrul-yazilim/etiket_olusturucu/main/guncelleme.txt"
        global APP_VERSION
        MEVCUT_SURUM = APP_VERSION.replace("v", "")  # Sadece rakamlı kısmı karşılaştırmak için
        try:
            resp = requests.get(GUNCELLEME_URL, timeout=5)
            if resp.status_code == 200:
                lines = resp.text.splitlines()
                guncel_surum = None
                indir_url = None
                changelog = ""
                for line in lines:
                    if line.startswith("version="):
                        guncel_surum = line.split("=", 1)[1].strip()
                    elif line.startswith("url="):
                        indir_url = line.split("=", 1)[1].strip()
                    elif line.startswith("changelog="):
                        changelog = line.split("=", 1)[1].strip()
                if guncel_surum and guncel_surum != MEVCUT_SURUM:
                    self._guncelleme_popup(guncel_surum, indir_url, changelog)
                else:
                    self.show_info("Güncelleme", "Güncellemeler Kontrol Edildi.Güncel Sürümü Kullanıyorsunuz.")
            else:
                self.show_error("Güncelleme", "Güncelleme sunucusuna ulaşılamadı.")
        except Exception as e:
            self.show_error("Güncelleme", f"Güncelleme kontrolü başarısız: {e}")

    def _guncelleme_popup(self, yeni_surum, indir_url, changelog):
        popup = tk.Toplevel(self.root)
        popup.title("Güncelleme Mevcut")
        popup.geometry("400x260")
        popup.configure(bg=BG_COLOR)
        msg = (
            f"Yeni sürüm bulundu: {yeni_surum}\n\n"
            f"Değişiklikler:\n{changelog}\n\n"
            "Lütfen açılan exe dosyasını indirip çalıştırınız.\n"
            "Eski sürümü bilgisayarınızdan kaldırabilirsiniz."
        )
        tk.Label(popup, text=msg, font=("Arial", 10), bg=BG_COLOR, wraplength=360, justify="left").pack(pady=14)
        def indir():
            import webbrowser
            webbrowser.open(indir_url)
            popup.destroy()
        btn1 = self.styled_button(popup, text="Evet, indir", command=indir, font=("Arial", 9, "bold"))
        btn1.pack(side=tk.LEFT, padx=18, pady=8)
        btn2 = self.styled_button(popup, text="Hayır, mevcut sürümü kullan", command=popup.destroy, font=("Arial", 9, "bold"), bg="#d32f2f", activebackground="#b71c1c")
        btn2.pack(side=tk.RIGHT, padx=18, pady=8)
        popup.transient(self.root)
        popup.grab_set()

    def download_kilavuz(self):
        from tkinter import filedialog, messagebox
        KULLANIM_KILAVUZU_METNI = """
SATRANÇ TURNUVASI KUPA/MADALYA ETİKETİ ve KATILIM BELGESİ OLUŞTURUCU - KULLANIM KILAVUZU
-------------------------------------------------------------------------------

1. BAŞLANGIÇ ve GENEL BİLGİLER
- Program, satranç turnuvaları için kupa/madalya etiketi ve katılım belgesi hazırlamanıza yardımcı olur.
- Menüden yeni bir turnuva başlatabilir, ödül ve kategori ayarlarını yapabilir, çıktı alabilirsiniz.

2. TURNUVA BİLGİLERİNİ GİRME
- Turnuva adı, ili, tarihi ve organizatör bilgilerini girin.
- Gerekirse turnuva logosu ekleyebilirsiniz (sol/sağ üst köşe için).

3. KATEGORİ ve ÖDÜL AYARLARI
- Kategorileri (ör. Genel, Kadınlar, 8 Yaş ve Altı vb.) ekleyin.
- Her kategori için ödül sıralaması ve metinlerini belirleyin (ör. 1., 2., 3. gibi).
- "Ödül Ekle" ve "Düzenle" butonları ile kategori/ödül bilgilerini güncelleyebilirsiniz.

4. ETİKET ve BELGE AYARLARI
- Etiket boyutu, yazı tipi, renk ve hizalama gibi ayarları yapabilirsiniz.
- Katılım belgesi için özel mesaj ekleyebilirsiniz.
- Gerekirse toplu katılım belgesi veya ödül etiketi çıktısı alabilirsiniz.

5. PDF OLUŞTURMA ve KAYDETME
- Tüm ayarları tamamladıktan sonra "PDF Oluştur" veya "Önizle" seçeneklerini kullanın.
- Oluşan PDF dosyasını istediğiniz klasöre kaydedebilirsiniz.
- Toplu çıktı almak için ilgili menüleri kullanın.

6. KULLANIM İPUÇLARI
- Menüdeki "Kullanım Kılavuzunu İndir" seçeneğiyle bu kılavuzu istediğiniz yere kaydedebilirsiniz.
- "Geri Bildirim" menüsünden öneri ve hata bildirimlerinizi iletebilirsiniz.
- Programı güncel tutmak için "Güncellemeleri Kontrol Et" seçeneğini kullanın.
- Kupa/madalya etiketleri ve belgeler için farklı yazı tipleri ve renkler deneyebilirsiniz.
- Kategorileri ve ödülleri kaydedip daha sonra tekrar yükleyebilirsiniz (isteğe bağlı).

7. SIK SORULAN SORULAR
- S: PDF dosyası açılmıyor veya yazı tipleri bozuk görünüyor. Ne yapmalıyım?
  C: Bilgisayarınızda gerekli fontlar yüklü olmalı ve PDF okuyucunuz güncel olmalı.
- S: Kendi logomu ekleyebilir miyim?
  C: Evet, logo seçme butonunu kullanarak PNG/JPG formatında logo ekleyebilirsiniz.
- S: Programı başka bilgisayarda kullanabilir miyim?
  C: Evet, .exe dosyasını ve gerekirse font klasörünü birlikte taşıyabilirsiniz.

8. TEKNİK DESTEK ve GÜNCELLEME
- Güncellemeler ve yeni sürümler için geliştiriciyle iletişime geçin veya GitHub sayfasını ziyaret edin.
- Hata ve önerileriniz için program içi geri bildirim formunu kullanabilirsiniz.

İYİ TURNUVALAR DİLERİZ!
"""
        try:
            hedef = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Metin Dosyası", "*.txt")],
                initialfile="KULLANIM_KILAVUZU.txt",
                title="Kılavuzu Kaydet"
            )
            if hedef:
                with open(hedef, "w", encoding="utf-8") as f:
                    f.write(KULLANIM_KILAVUZU_METNI)
                messagebox.showinfo("Kılavuz Kaydedildi", f"Kullanım kılavuzu başarıyla kaydedildi:\n{hedef}")
        except Exception as e:
            messagebox.showerror("Hata", f"Kılavuz kaydedilemedi: {e}")

def katilim_belgesi_pdf(pdf_path, isim_kategori_listesi, il, turnuva_adi, tarih, logo_sol_path, logo_sag_path, direktoru, bashakemi, mesaj):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    import textwrap
    FONT_PATH = resource_path(os.path.join("ttf", "DejaVuSans.ttf"))
    pdfmetrics.registerFont(TTFont("DejaVu", FONT_PATH))
    c = canvas.Canvas(pdf_path, pagesize=landscape(A4))
    w, h = landscape(A4)
    left_margin = 1.0*cm
    right_margin = 1.0*cm
    top_margin = 1.0*cm
    bottom_margin = 2.5*cm
    max_text_width = w - left_margin - right_margin
    def draw_wrapped_centered(text, font, size, y, max_width, line_space=1.25):
        c.setFont(font, size)
        wrapper = textwrap.TextWrapper(width=90)
        lines = []
        for t in text.split("\n"):
            lines.extend(wrapper.wrap(t) if len(t) > 0 else [""])
        for i, line in enumerate(lines):
            c.drawCentredString(w/2, y - i*size*line_space, line)
        return y - len(lines)*size*line_space
    for isim, kategori in isim_kategori_listesi:
        # Logoları üst köşelere daha yakın yerleştir
        logo_y = h - 1.0*cm - 3*cm  # 1cm üst boşluk
        sol_logo = resource_path(logo_sol_path) if logo_sol_path else ""
        sag_logo = resource_path(logo_sag_path) if logo_sag_path else ""
        if sol_logo and os.path.exists(sol_logo):
            c.drawImage(ImageReader(sol_logo), left_margin, logo_y, width=3*cm, height=3*cm, mask='auto')
        if sag_logo and os.path.exists(sag_logo):
            c.drawImage(ImageReader(sag_logo), w-right_margin-3*cm, logo_y, width=3*cm, height=3*cm, mask='auto')
        # Metinleri biraz daha aşağıdan başlat
        y = h - 1.0*cm - 3.5*cm  # logoların altından biraz daha aşağı
        # Başlık
        c.setFont("DejaVu", 36)
        c.setFillColorRGB(0.13, 0.22, 0.45)
        c.drawCentredString(w/2, y, "KATILIM BELGESİ")
        y -= 2.2*cm
        # Sayın Adı Soyadı
        c.setFont("DejaVu", 22)
        c.setFillColorRGB(0,0,0)
        c.drawCentredString(w/2, y, f"Sayın {isim}")
        y -= 1.1*cm
        # Kategori
        c.setFont("DejaVu", 18)
        c.setFillColorRGB(0.15,0.15,0.15)
        c.drawCentredString(w/2, y, f"Kategori: {kategori}")
        y -= 1.5*cm
        # Açıklama metni (tamamı ortada ve wrap'li)
        c.setFont("DejaVu", 15)
        c.setFillColorRGB(0,0,0)
        aciklama = f"{il} ilinde, {tarih} tarihleri arasında düzenlenen {turnuva_adi}'na göstermiş olduğunuz değerli katılımınızdan dolayı teşekkür ederiz."
        y = draw_wrapped_centered(aciklama, "DejaVu", 15, y, max_text_width, line_space=1.25)
        y -= 0.7*cm
        # Teşekkür ve başarılar metni (kullanıcıdan alınan mesaj, wrap'li ve ortalanmış)
        if mesaj:
            y = draw_wrapped_centered(mesaj, "DejaVu", 15, y, max_text_width, line_space=1.25)
        # Alt köşeler: imzalar
        imza_y = bottom_margin + 2.2*cm
        isim_y = bottom_margin + 1.2*cm
        # Sol alt: Direktör
        c.setFont("DejaVu", 13)
        c.drawCentredString(left_margin+4*cm, isim_y, direktoru + " (Turnuva Direktörü)")
        # Sağ alt: Başhakem
        c.drawCentredString(w-right_margin-4*cm, isim_y, bashakemi + " (Başhakem)")
        c.showPage()
    c.save()



def TarihSecici(master, var, baslangic=None):
    top = tk.Toplevel(master)
    top.title("Tarih Seç")
    cal = Calendar(top, selectmode='day', locale='tr_TR')
    cal.pack(padx=10, pady=10)
    def sec():
        sec_date = cal.selection_get()
        if sec_date:
            var.set(sec_date.strftime("%d.%m.%Y"))
            top.destroy()
    tk.Button(top, text="Seç", command=sec).pack(pady=5)
    if baslangic and baslangic != "Seçiniz":
        try:
            d, m, y = map(int, baslangic.split("."))
            cal.selection_set(datetime.date(y, m, d))
        except:
            pass
    cal.bind("<Double-1>", lambda e: sec())
    top.bind('<Escape>', lambda e: top.destroy())

APP_VERSION = "v22.07.24.02"

def start_app():
    save_kilavuz_on_start()
    root = tk.Tk()
    root.withdraw()  # Ana pencereyi gizle
    def show_main():
        root.deiconify()
        app = EtiketUygulamasi(root)
    splash_path = resource_path("baslangic.jpg")
    splash_screen_then_start(root, show_main, splash_path, splash_time=4000)
    root.mainloop()

def save_kilavuz_on_start():
    import shutil
    import os
    try:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        hedef = os.path.join(desktop, "KULLANIM_KILAVUZU.txt")
        if not os.path.exists(hedef):
            shutil.copy("KULLANIM_KILAVUZU.txt", hedef)
    except Exception:
        pass

# Ana başlatıcı fonksiyon
if __name__ == "__main__":
    start_app() 